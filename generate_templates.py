"""
generate_templates.py
Domestic / International 기본 Expense Report Excel 템플릿 생성기
- 관리자 페이지에서 "Generate Default Templates" 버튼 클릭 시 호출
- 또는 독립 실행: python generate_templates.py
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
import os, sys
from pathlib import Path

# ── 공통 스타일 상수 ─────────────────────────────────────
C_HEADER_BG   = "1E3A5F"   # 진한 네이비 (헤더)
C_HEADER_FG   = "FFFFFF"
C_SUBHDR_BG   = "2E6DA4"   # 중간 블루 (서브헤더)
C_SUBHDR_FG   = "FFFFFF"
C_LABEL_BG    = "D9E1F2"   # 연한 파랑 (레이블)
C_TOTAL_BG    = "FFF2CC"   # 연한 노랑 (합계)
C_SIGN_BG     = "F2F2F2"   # 연한 회색 (서명)
C_BORDER      = "AAAAAA"
C_ALT_ROW     = "EEF3FB"   # 짝수 행 배경

THIN  = Side(style="thin",   color=C_BORDER)
THICK = Side(style="medium", color="888888")

def border(left=None, right=None, top=None, bottom=None):
    return Border(left=left or THIN, right=right or THIN,
                  top=top or THIN,   bottom=bottom or THIN)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def configure_print(ws, area: str, landscape: bool = False):
    ws.print_area = area
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

def merge_write(ws, cell_range, value, fnt=None, aln=None, brd=None, fll=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if fnt: cell.font = fnt
    if aln: cell.alignment = aln
    if brd: cell.border = brd
    if fll: cell.fill = fll

def label_row(ws, row, label, col_start=1, col_end=3, value_col_end=7):
    """레이블 + 입력칸 한 줄"""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    lbl = ws.cell(row=row, column=col_start)
    lbl.value     = label
    lbl.font      = font(bold=True, size=9)
    lbl.fill      = fill(C_LABEL_BG)
    lbl.alignment = align("left")
    lbl.border    = border()
    ws.merge_cells(start_row=row, start_column=col_end+1,
                   end_row=row,   end_column=value_col_end)
    val = ws.cell(row=row, column=col_end+1)
    val.border    = border()
    val.alignment = align("left")
    return val   # 값 셀 반환


# ═══════════════════════════════════════════════════
# DOMESTIC TEMPLATE
# ═══════════════════════════════════════════════════

def create_domestic_template(output_path: str, cfg: dict = None):
    """
    Domestic Expense Report (USD only)
    Sheets: List  |  Report (template for weekly copies)
    """
    cfg = cfg or {}
    company = cfg.get("company", {})
    types   = [t for t in cfg.get("expense_types", _default_types()) if t.get("enabled")]
    fields  = cfg.get("fields", _default_fields())

    wb = Workbook()

    # ── LIST sheet ──────────────────────────────────────
    ws_list = wb.active
    ws_list.title = "List"
    _build_list_sheet_domestic(ws_list, company, fields)

    # ── REPORT (weekly template) sheet ─────────────────
    ws_rpt = wb.create_sheet("Report")
    _build_report_sheet_domestic(ws_rpt, company, fields, types)

    wb.save(output_path)
    return output_path


def _build_list_sheet_domestic(ws, company, fields):
    """List 시트 — 전체 영수증 목록 (USD only)"""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 16

    # 헤더 배너
    merge_write(ws, "A1:I1",
        company.get("name") or "EXPENSE REPORT",
        fnt=font(bold=True, size=16, color=C_HEADER_FG),
        aln=align("center"),
        fll=fill(C_HEADER_BG))
    merge_write(ws, "A2:I2", "Itemized Receipt List",
        fnt=font(size=9, color="888888"),
        aln=align("center"))

    # 직원 정보 행
    info_row = 4
    ws.row_dimensions[info_row].height = 18
    _write_info_section_domestic(ws, info_row, fields)

    # 컬럼 헤더 (행 7)
    hdr_row = 7
    ws.row_dimensions[hdr_row].height = 20
    headers = ["#", "Date", "Merchant / Description", "Category",
               "Amount (USD)", "Memo"]
    widths  = [4,   12,     36,                        22,
               14,           28]
    cols    = ["A", "B", "C", "D", "E", "F"]
    for col, hdr, w in zip(cols, headers, widths):
        c = ws[f"{col}{hdr_row}"]
        c.value     = hdr
        c.font      = font(bold=True, size=9, color=C_HEADER_FG)
        c.fill      = fill(C_SUBHDR_BG)
        c.alignment = align("center")
        c.border    = border()
        set_col_width(ws, col, w)

    # 데이터 행 자리 (행 8 ~ 57, 50개)
    for i in range(1, 51):
        r = hdr_row + i
        ws.row_dimensions[r].height = 16
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        for col_idx, col in enumerate(["A","B","C","D","E","F"], 1):
            c = ws.cell(row=r, column=col_idx)
            c.fill      = fill(bg)
            c.border    = border()
            c.alignment = align("center" if col in ("A","B","E") else "left")
            if col == "A": c.value = i
            if col == "E": c.number_format = '#,##0.00'

    # 합계 행
    total_row = hdr_row + 51
    ws.row_dimensions[total_row].height = 20
    for col_idx, col in enumerate(["A","B","C","D","E","F"], 1):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = fill(C_TOTAL_BG)
        c.border = border(top=THICK, bottom=THICK)
        c.font   = font(bold=True, size=9)
    ws.cell(row=total_row, column=4).value = "TOTAL"
    ws.cell(row=total_row, column=4).alignment = align("right")
    ws.cell(row=total_row, column=5).value = f"=SUM(E{hdr_row+1}:E{hdr_row+50})"
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws.cell(row=total_row, column=5).alignment = align("center")

    # 서명란
    _write_signature_section(ws, total_row + 2, col_count=6)
    configure_print(ws, f"A1:F{total_row+4}")


def _write_info_section_domestic(ws, start_row, fields):
    enabled = {k: v for k, v in fields.items() if v.get("enabled")}
    row = start_row
    pairs = [
        ("employee_name", "period"),
        ("department",    "manager"),
        ("employee_id",   "project"),
    ]
    for left_key, right_key in pairs:
        if left_key in enabled or right_key in enabled:
            ws.row_dimensions[row].height = 18
            if left_key in enabled:
                _info_cell_pair(ws, row, 1, enabled[left_key]["label"])
            if right_key in enabled:
                _info_cell_pair(ws, row, 5, enabled[right_key]["label"])
            row += 1
    return row


def _info_cell_pair(ws, row, col_start, label):
    lbl = ws.cell(row=row, column=col_start)
    lbl.value     = label
    lbl.font      = font(bold=True, size=9)
    lbl.fill      = fill(C_LABEL_BG)
    lbl.alignment = align("left")
    lbl.border    = border()
    val = ws.cell(row=row, column=col_start + 1)
    val.border    = border()
    ws.merge_cells(start_row=row, start_column=col_start+1,
                   end_row=row,   end_column=col_start+2)


def _write_signature_section(ws, start_row, col_count=9):
    row = start_row
    ws.row_dimensions[row].height   = 14
    ws.row_dimensions[row+1].height = 28
    ws.row_dimensions[row+2].height = 14

    sign_labels = ["Employee Signature", "Supervisor Approval", "Finance Approval"]
    chunk = max(col_count // len(sign_labels), 2)
    for i, lbl in enumerate(sign_labels):
        c1 = i * chunk + 1
        c2 = c1 + chunk - 1
        col_letter = get_column_letter(c1)
        col_letter2 = get_column_letter(c2)
        ws.merge_cells(f"{col_letter}{row}:{col_letter2}{row}")
        ws.merge_cells(f"{col_letter}{row+1}:{col_letter2}{row+1}")
        ws.merge_cells(f"{col_letter}{row+2}:{col_letter2}{row+2}")

        lbl_cell = ws[f"{col_letter}{row}"]
        lbl_cell.value     = lbl
        lbl_cell.font      = font(bold=True, size=8)
        lbl_cell.fill      = fill(C_SIGN_BG)
        lbl_cell.alignment = align("center")
        lbl_cell.border    = border()

        sign_cell = ws[f"{col_letter}{row+1}"]
        sign_cell.fill   = fill("FFFFFF")
        sign_cell.border = border()

        date_cell = ws[f"{col_letter}{row+2}"]
        date_cell.value     = "Date: _______________"
        date_cell.font      = font(size=8, color="888888")
        date_cell.fill      = fill(C_SIGN_BG)
        date_cell.alignment = align("center")
        date_cell.border    = border()


def _build_report_sheet_domestic(ws, company, fields, types):
    """
    Report 시트 (주간 템플릿)
    행 구조:
      1-2 : 헤더 배너
      3   : 공백
      4-6 : 직원 정보
      7   : 공백
      8   : 날짜 헤더 (D~J = Mon~Sun)
      9.. : Expense type 행들
      N   : Daily Total
      N+1 : Weekly Total
      N+3 : 서명란
    """
    ws.sheet_view.showGridLines = False

    # ── 배너 ──
    ws.row_dimensions[1].height = 36
    merge_write(ws, "A1:K1",
        company.get("name") or "EXPENSE REPORT",
        fnt=font(bold=True, size=14, color=C_HEADER_FG),
        aln=align("center"),
        fll=fill(C_HEADER_BG))
    ws.row_dimensions[2].height = 14
    merge_write(ws, "A2:K2", "Weekly Expense Report — Domestic",
        fnt=font(size=9, color="888888"),
        aln=align("center"))

    # ── 직원 정보 ──
    ws.row_dimensions[3].height = 8
    _write_info_section_report(ws, 4, fields, col_span=11)

    # ── 날짜 헤더 행 (row 8) ──
    DATE_ROW = 8
    ws.row_dimensions[DATE_ROW].height = 20
    day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    # col A=1:category label, B=2:empty, C=3:label(USD), D~J = 7 days, K=11:weekly total
    ws.cell(row=DATE_ROW, column=1).value = "Expense Category"
    for c in range(1, 4):
        cell = ws.cell(row=DATE_ROW, column=c)
        cell.font  = font(bold=True, size=9, color=C_HEADER_FG)
        cell.fill  = fill(C_SUBHDR_BG)
        cell.alignment = align("center")
        cell.border = border()
    for i, day in enumerate(day_headers):
        col = 4 + i   # D=4 .. J=10
        cell = ws.cell(row=DATE_ROW, column=col)
        cell.value = day          # 날짜는 auto_create_weekly_sheets 가 채움
        cell.font  = font(bold=True, size=9, color=C_HEADER_FG)
        cell.fill  = fill(C_SUBHDR_BG)
        cell.alignment = align("center")
        cell.border = border()
    # Weekly Total 헤더
    wt = ws.cell(row=DATE_ROW, column=11)
    wt.value = "Weekly\nTotal"
    wt.font  = font(bold=True, size=9, color=C_HEADER_FG)
    wt.fill  = fill(C_HEADER_BG)
    wt.alignment = align("center", wrap=True)
    wt.border = border()

    # 열 너비
    set_col_width(ws, "A", 24)
    set_col_width(ws, "B", 3)
    set_col_width(ws, "C", 3)
    for col_letter in ["D","E","F","G","H","I","J"]:
        set_col_width(ws, col_letter, 11)
    set_col_width(ws, "K", 13)

    # ── Expense Type 행들 ──
    data_start = DATE_ROW + 1
    for idx, t in enumerate(types):
        r = data_start + idx
        ws.row_dimensions[r].height = 17
        bg = C_ALT_ROW if idx % 2 == 0 else "FFFFFF"

        # 카테고리 레이블 (A~C 병합)
        ws.merge_cells(f"A{r}:C{r}")
        lbl = ws[f"A{r}"]
        lbl.value     = t["label"]
        lbl.font      = font(size=9, bold=False)
        lbl.fill      = fill(C_LABEL_BG)
        lbl.alignment = align("left")
        lbl.border    = border()

        # 날짜별 금액 셀 D~J
        day_cells = []
        for col in range(4, 11):
            c = ws.cell(row=r, column=col)
            c.fill          = fill(bg)
            c.border        = border()
            c.alignment     = align("center")
            c.number_format = '#,##0.00'
            day_cells.append(get_column_letter(col))

        # Weekly Total (K) = SUM(D:J)
        k = ws.cell(row=r, column=11)
        k.value          = f"=SUM(D{r}:J{r})"
        k.font           = font(bold=True, size=9)
        k.fill           = fill(C_TOTAL_BG)
        k.border         = border()
        k.alignment      = align("center")
        k.number_format  = '#,##0.00'

    # ── Daily Total 행 ──
    total_row = data_start + len(types)
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:C{total_row}")
    dt_lbl = ws[f"A{total_row}"]
    dt_lbl.value     = "DAILY TOTAL"
    dt_lbl.font      = font(bold=True, size=9, color=C_HEADER_FG)
    dt_lbl.fill      = fill(C_SUBHDR_BG)
    dt_lbl.alignment = align("center")
    dt_lbl.border    = border(top=THICK, bottom=THICK)

    for col in range(4, 11):
        col_l = get_column_letter(col)
        c = ws.cell(row=total_row, column=col)
        c.value         = f"=SUM({col_l}{data_start}:{col_l}{total_row-1})"
        c.font          = font(bold=True, size=9)
        c.fill          = fill(C_TOTAL_BG)
        c.border        = border(top=THICK, bottom=THICK)
        c.alignment     = align("center")
        c.number_format = '#,##0.00'

    # Weekly Grand Total (K)
    k = ws.cell(row=total_row, column=11)
    k.value         = f"=SUM(K{data_start}:K{total_row-1})"
    k.font          = font(bold=True, size=11)
    k.fill          = fill(C_TOTAL_BG)
    k.border        = border(top=THICK, bottom=THICK)
    k.alignment     = align("center")
    k.number_format = '#,##0.00'

    # ── 서명란 ──
    _write_signature_section(ws, total_row + 2, col_count=11)
    configure_print(ws, f"A1:K{total_row+4}", landscape=True)


def _write_info_section_report(ws, start_row, fields, col_span=11):
    enabled = {k: v for k, v in fields.items() if v.get("enabled")}
    row = start_row
    pairs = [
        ("employee_name", "period"),
        ("department",    "manager"),
        ("employee_id",   "project"),
    ]
    half = col_span // 2
    for left_key, right_key in pairs:
        if left_key in enabled or right_key in enabled:
            ws.row_dimensions[row].height = 18
            if left_key in enabled:
                _info_report_pair(ws, row, 1, half - 1, enabled[left_key]["label"])
            if right_key in enabled:
                _info_report_pair(ws, row, half + 1, col_span, enabled[right_key]["label"])
            row += 1
    return row


def _info_report_pair(ws, row, c1, c2, label):
    mid = (c1 + c2) // 2
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=mid - 1)
    lbl = ws.cell(row=row, column=c1)
    lbl.value     = label
    lbl.font      = font(bold=True, size=9)
    lbl.fill      = fill(C_LABEL_BG)
    lbl.alignment = align("left")
    lbl.border    = border()

    ws.merge_cells(start_row=row, start_column=mid, end_row=row, end_column=c2)
    val = ws.cell(row=row, column=mid)
    val.border    = border()
    val.alignment = align("left")


# ═══════════════════════════════════════════════════
# INTERNATIONAL TEMPLATE
# ═══════════════════════════════════════════════════

def create_international_template(output_path: str, cfg: dict = None):
    """
    International Expense Report
    - 외화 금액 + 환율 + USD 환산
    Sheets: List  |  Report (weekly template)
    """
    cfg = cfg or {}
    company = cfg.get("company", {})
    types   = [t for t in cfg.get("expense_types", _default_types()) if t.get("enabled")]
    fields  = cfg.get("fields", _default_fields())

    wb = Workbook()

    ws_list = wb.active
    ws_list.title = "List"
    _build_list_sheet_intl(ws_list, company, fields)

    ws_rpt = wb.create_sheet("Report")
    _build_report_sheet_intl(ws_rpt, company, fields, types)

    wb.save(output_path)
    return output_path


def _build_list_sheet_intl(ws, company, fields):
    """List 시트 — 외화 + USD 병기"""
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 16

    merge_write(ws, "A1:K1",
        company.get("name") or "EXPENSE REPORT",
        fnt=font(bold=True, size=16, color=C_HEADER_FG),
        aln=align("center"),
        fll=fill(C_HEADER_BG))
    merge_write(ws, "A2:K2", "Itemized Receipt List — International",
        fnt=font(size=9, color="888888"),
        aln=align("center"))

    info_row = 4
    _write_info_section_domestic(ws, info_row, fields)

    # Trip destination 추가 행
    ws.row_dimensions[info_row + 3].height = 18
    _info_cell_pair(ws, info_row + 3, 1, "Trip Destination")
    _info_cell_pair(ws, info_row + 3, 5, "Purpose of Travel")

    hdr_row = 8
    ws.row_dimensions[hdr_row].height = 20
    headers = ["#", "Date", "Merchant / Description", "Category",
               "Currency", "Foreign Amt", "Exch. Rate", "Amount (USD)", "Memo"]
    widths  = [4,   12,     28,                        22,
               9,         13,         11,            14,           24]
    for col_idx, (hdr, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=hdr_row, column=col_idx)
        c.value     = hdr
        c.font      = font(bold=True, size=9, color=C_HEADER_FG)
        c.fill      = fill(C_SUBHDR_BG)
        c.alignment = align("center")
        c.border    = border()
        set_col_width(ws, get_column_letter(col_idx), w)

    for i in range(1, 51):
        r = hdr_row + i
        ws.row_dimensions[r].height = 16
        bg = C_ALT_ROW if i % 2 == 0 else "FFFFFF"
        for col_idx in range(1, 10):
            c = ws.cell(row=r, column=col_idx)
            c.fill   = fill(bg)
            c.border = border()
            c.alignment = align("center" if col_idx in (1, 2, 5, 6, 7, 8) else "left")
        ws.cell(row=r, column=1).value = i
        ws.cell(row=r, column=8).number_format = '#,##0.00'
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=7).number_format = '#,##0.0000'

    total_row = hdr_row + 51
    ws.row_dimensions[total_row].height = 20
    for col_idx in range(1, 10):
        c = ws.cell(row=total_row, column=col_idx)
        c.fill   = fill(C_TOTAL_BG)
        c.border = border(top=THICK, bottom=THICK)
        c.font   = font(bold=True, size=9)
    ws.cell(row=total_row, column=7).value = "TOTAL (USD)"
    ws.cell(row=total_row, column=7).alignment = align("right")
    ws.cell(row=total_row, column=8).value = f"=SUM(H{hdr_row+1}:H{hdr_row+50})"
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'

    _write_signature_section(ws, total_row + 2, col_count=9)
    configure_print(ws, f"A1:I{total_row+4}", landscape=True)


def _build_report_sheet_intl(ws, company, fields, types):
    """
    Report 시트 (국제 주간 템플릿)
    각 날짜 열이 두 개 (Foreign | USD)로 구성
    """
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 36
    merge_write(ws, "A1:P1",
        company.get("name") or "EXPENSE REPORT",
        fnt=font(bold=True, size=14, color=C_HEADER_FG),
        aln=align("center"),
        fll=fill(C_HEADER_BG))
    ws.row_dimensions[2].height = 14
    merge_write(ws, "A2:P2", "Weekly Expense Report — International",
        fnt=font(size=9, color="888888"),
        aln=align("center"))
    ws.row_dimensions[3].height = 8

    _write_info_section_report(ws, 4, fields, col_span=16)

    # 국제선 추가 정보 행
    ws.row_dimensions[7].height = 18
    _info_report_pair(ws, 7, 1, 8, "Destination Country")
    _info_report_pair(ws, 7, 9, 16, "Local Currency")

    # 날짜 헤더 (row 8 = merged date, row 9 = day group, row 10 = sub-header)
    DATE_ROW = 8
    GRP_ROW = 9
    SUB_ROW = 10
    ws.row_dimensions[DATE_ROW].height = 18
    ws.row_dimensions[GRP_ROW].height = 18
    ws.row_dimensions[SUB_ROW].height = 16

    # col A = category label (A~B 병합)
    ws.merge_cells(f"A{DATE_ROW}:B{DATE_ROW}")
    ws.merge_cells(f"A{GRP_ROW}:B{GRP_ROW}")
    ws.merge_cells(f"A{SUB_ROW}:B{SUB_ROW}")
    for rr in (DATE_ROW, GRP_ROW, SUB_ROW):
        c = ws[f"A{rr}"]
        c.value     = "Expense Category" if rr == GRP_ROW else ""
        c.font      = font(bold=True, size=9, color=C_HEADER_FG)
        c.fill      = fill(C_SUBHDR_BG)
        c.alignment = align("center")
        c.border    = border()

    # 7일 × 2열 (Foreign, USD) = 14열 (C~P)
    day_labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for i, day in enumerate(day_labels):
        g_col   = 3 + i * 2      # C, E, G, I, K, M, O
        usd_col = g_col + 1      # D, F, H, J, L, N, P
        g_letter   = get_column_letter(g_col)
        usd_letter = get_column_letter(usd_col)

        # 날짜 헤더 (auto_create_weekly_sheets 가 채움)
        ws.merge_cells(f"{g_letter}{DATE_ROW}:{usd_letter}{DATE_ROW}")
        dc = ws[f"{g_letter}{DATE_ROW}"]
        dc.font      = font(bold=True, size=9)
        dc.alignment = align("center")
        dc.border    = border()

        # 그룹 헤더 (요일)
        ws.merge_cells(f"{g_letter}{GRP_ROW}:{usd_letter}{GRP_ROW}")
        gc = ws[f"{g_letter}{GRP_ROW}"]
        gc.value     = day
        gc.font      = font(bold=True, size=9, color=C_HEADER_FG)
        gc.fill      = fill(C_SUBHDR_BG)
        gc.alignment = align("center")
        gc.border    = border()

        # 서브 헤더
        for col, lbl in [(g_col, "Foreign"), (usd_col, "USD")]:
            c = ws.cell(row=SUB_ROW, column=col)
            c.value     = lbl
            c.font      = font(bold=True, size=8, color=C_HEADER_FG)
            c.fill      = fill(C_HEADER_BG)
            c.alignment = align("center")
            c.border    = border()
            set_col_width(ws, get_column_letter(col), 10)

    set_col_width(ws, "A", 24)
    set_col_width(ws, "B", 3)

    # Weekly Total 열 (Q=17)
    WTOTAL_COL = 3 + 7 * 2  # = 17
    for rr in (GRP_ROW, SUB_ROW):
        wt = ws.cell(row=rr, column=WTOTAL_COL)
        wt.value     = "Weekly\nTotal (USD)" if rr == GRP_ROW else ""
        wt.font      = font(bold=True, size=9, color=C_HEADER_FG)
        wt.fill      = fill(C_HEADER_BG)
        wt.alignment = align("center", wrap=True)
        wt.border    = border()
    set_col_width(ws, get_column_letter(WTOTAL_COL), 14)

    # Expense type 행들
    data_start = SUB_ROW + 1
    usd_cols = [get_column_letter(3 + i * 2 + 1) for i in range(7)]

    for idx, t in enumerate(types):
        r  = data_start + idx
        bg = C_ALT_ROW if idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r].height = 17

        ws.merge_cells(f"A{r}:B{r}")
        lbl = ws[f"A{r}"]
        lbl.value     = t["label"]
        lbl.font      = font(size=9)
        lbl.fill      = fill(C_LABEL_BG)
        lbl.alignment = align("left")
        lbl.border    = border()

        for i in range(7):
            g_col   = 3 + i * 2
            usd_col = g_col + 1
            for col in (g_col, usd_col):
                c = ws.cell(row=r, column=col)
                c.fill          = fill(bg)
                c.border        = border()
                c.alignment     = align("center")
                c.number_format = '#,##0.00'

        # Weekly total
        usd_sum = "+".join(f"{u}{r}" for u in usd_cols)
        k = ws.cell(row=r, column=WTOTAL_COL)
        k.value         = f"={usd_sum}"
        k.font          = font(bold=True, size=9)
        k.fill          = fill(C_TOTAL_BG)
        k.border        = border()
        k.alignment     = align("center")
        k.number_format = '#,##0.00'

    # Daily Total
    total_row = data_start + len(types)
    ws.row_dimensions[total_row].height = 20
    ws.merge_cells(f"A{total_row}:B{total_row}")
    dt = ws[f"A{total_row}"]
    dt.value     = "DAILY TOTAL (USD)"
    dt.font      = font(bold=True, size=9, color=C_HEADER_FG)
    dt.fill      = fill(C_SUBHDR_BG)
    dt.alignment = align("center")
    dt.border    = border(top=THICK, bottom=THICK)

    for i in range(7):
        for is_usd, col in enumerate([3 + i*2, 3 + i*2 + 1]):
            col_l = get_column_letter(col)
            c = ws.cell(row=total_row, column=col)
            if is_usd:
                c.value = f"=SUM({col_l}{data_start}:{col_l}{total_row-1})"
                c.font  = font(bold=True, size=9)
            c.fill          = fill(C_TOTAL_BG)
            c.border        = border(top=THICK, bottom=THICK)
            c.alignment     = align("center")
            c.number_format = '#,##0.00'

    # Grand Weekly Total
    wt = ws.cell(row=total_row, column=WTOTAL_COL)
    usd_sum2 = "+".join(f"{u}{total_row}" for u in usd_cols)
    wt.value         = f"={usd_sum2}"
    wt.font          = font(bold=True, size=11)
    wt.fill          = fill(C_TOTAL_BG)
    wt.border        = border(top=THICK, bottom=THICK)
    wt.alignment     = align("center")
    wt.number_format = '#,##0.00'

    _write_signature_section(ws, total_row + 2, col_count=WTOTAL_COL)
    configure_print(ws, f"A1:Q{total_row+4}", landscape=True)


# ═══════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════

def _default_types():
    return [
        {"id": "BREAKFAST",     "label": "Breakfast",         "enabled": True},
        {"id": "LUNCH",         "label": "Lunch",             "enabled": True},
        {"id": "DINNER",        "label": "Dinner",            "enabled": True},
        {"id": "ENTERTAINMENT", "label": "Entertainment",     "enabled": True},
        {"id": "LODGING",       "label": "Lodging / Hotel",   "enabled": True},
        {"id": "AIRFARE",       "label": "Airfare",           "enabled": True},
        {"id": "CAR_RENTAL",    "label": "Car Rental",        "enabled": True},
        {"id": "TAXI",          "label": "Taxi / Rideshare",  "enabled": True},
        {"id": "MILEAGE",       "label": "Mileage",           "enabled": True},
        {"id": "PARKING",       "label": "Parking / Tolls",   "enabled": True},
        {"id": "PHONE",         "label": "Phone / Internet",  "enabled": True},
        {"id": "MISCELLANEOUS", "label": "Miscellaneous",     "enabled": True},
    ]

def _default_fields():
    return {
        "employee_name": {"enabled": True,  "label": "Employee Name"},
        "department":    {"enabled": True,  "label": "Department"},
        "employee_id":   {"enabled": False, "label": "Employee ID"},
        "manager":       {"enabled": False, "label": "Manager / Supervisor"},
        "project":       {"enabled": False, "label": "Project / Cost Center"},
        "period":        {"enabled": True,  "label": "Expense Period"},
    }


# ═══════════════════════════════════════════════════
# CLI 실행
# ═══════════════════════════════════════════════════

if __name__ == "__main__":
    import json
    out_dir = Path(__file__).parent / "default_template"
    out_dir.mkdir(exist_ok=True)

    cfg = {}
    cfg_path = Path(__file__).parent / "config.json"
    if cfg_path.exists():
        with open(cfg_path, encoding="utf-8") as f:
            cfg = json.load(f)

    dom_path = str(out_dir / "template_domestic.xlsx")
    intl_path = str(out_dir / "template_international.xlsx")

    create_domestic_template(dom_path, cfg)
    print(f"✓ Domestic template → {dom_path}")

    create_international_template(intl_path, cfg)
    print(f"✓ International template → {intl_path}")
