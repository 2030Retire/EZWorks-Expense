"""
Expense Report Processor — Web Application
영수증 이미지 → Claude API OCR → Excel 경비보고서 자동 생성
v2: Domestic / International 모드, Admin 설정 관리
"""
import os
import json
import copy
import csv
import io
import hashlib
from datetime import datetime, date as date_type, timedelta, timezone
from pathlib import Path
import uuid
import shutil
from urllib import request as urlrequest
from urllib import parse as urlparse
from urllib import error as urlerror

from user_db import (
    init_user_db,
    upsert_user_from_oidc,
    get_user_by_id,
    list_users,
    update_user_flags,
    update_user_profile,
    upsert_user_by_email,
    get_user_host_permissions,
    set_user_host_permissions,
    list_user_host_permissions,
    list_inbox_filter_presets,
    upsert_inbox_filter_preset,
    delete_inbox_filter_preset,
    list_shared_inbox_filter_presets,
    get_shared_inbox_filter_preset,
    upsert_shared_inbox_filter_preset,
    delete_shared_inbox_filter_preset,
    create_shared_inbox_filter_preset_audit_log,
    list_shared_inbox_filter_preset_audit_logs,
    get_user_wizard_preference,
    upsert_user_wizard_preference,
)
from receipt_inbox_db import (
    init_receipt_db,
    create_receipt,
    get_receipt,
    delete_receipt,
    find_receipt_by_hash,
    list_receipts,
    list_receipts_for_report,
    find_duplicate_receipt,
    update_receipt,
    create_report,
    get_report,
    list_reports,
    list_receipts_by_report_id,
    get_report_by_output,
    upsert_category_mapping,
    list_category_mappings,
    upsert_vendor_mapping,
    list_vendor_mappings,
    create_receipt_audit_log,
    list_receipt_audit_logs,
    list_receipt_audit_logs_by_host,
)

# ─────────────────────────────────────────────
# .env 로드 (python-dotenv 없이 동작)
# ─────────────────────────────────────────────
def _load_dotenv():
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    try:
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, _, val = line.partition("=")
                key, val = key.strip(), val.strip()
                if len(val) >= 2 and val[0] == val[-1] and val[0] in ('"', "'"):
                    val = val[1:-1]
                if key and val:
                    os.environ[key] = val
    except Exception:
        pass

_load_dotenv()

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, Response
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import Workbook, load_workbook
try:
    from authlib.integrations.flask_client import OAuth
except Exception:
    OAuth = None

from ocr import process_receipt_image
from excel_filler import fill_expense_report, DEFAULT_EXCHANGE_RATE, parse_date

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.secret_key = os.environ.get("SECRET_KEY", "change-this-in-production")
app.permanent_session_lifetime = timedelta(
    hours=int(os.environ.get("ADMIN_SESSION_HOURS", "12"))
)

BASE_DIR              = Path(__file__).parent
UPLOAD_FOLDER         = BASE_DIR / "uploads"
INBOX_UPLOAD_FOLDER   = UPLOAD_FOLDER / "inbox"
OUTPUT_FOLDER         = BASE_DIR / "outputs"
DEFAULT_TEMPLATE_DIR  = BASE_DIR / "default_template"
TENANT_TEMPLATE_ROOT  = BASE_DIR / "tenant_templates"
LOGO_FOLDER           = BASE_DIR / "static" / "logos"
LOGIN_BG_FOLDER       = BASE_DIR / "static" / "login_bg"
CONFIG_PATH           = BASE_DIR / "config.json"
CONFIG_DIR            = BASE_DIR / "configs"
TENANT_REGISTRY_PATH  = BASE_DIR / "tenants.json"

ALLOWED_IMAGE_EXT  = {"jpg", "jpeg", "png", "heic", "webp"}
ALLOWED_EXCEL_EXT  = {"xlsx", "xls"}
ALLOWED_LOGO_EXT   = {"png", "jpg", "jpeg", "gif", "svg", "webp"}

for folder in [UPLOAD_FOLDER, INBOX_UPLOAD_FOLDER, OUTPUT_FOLDER, DEFAULT_TEMPLATE_DIR, TENANT_TEMPLATE_ROOT, LOGO_FOLDER, LOGIN_BG_FOLDER, CONFIG_DIR]:
    folder.mkdir(parents=True, exist_ok=True)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
USER_DB_PATH = BASE_DIR / "users.db"
RECEIPT_DB_PATH = BASE_DIR / "receipt_inbox.db"

SSO_ENABLED = os.environ.get("SSO_ENABLED", "false").lower() in ("1", "true", "yes", "on")
SSO_PROVIDER = os.environ.get("SSO_PROVIDER", "microsoft").strip().lower()
MS_TENANT_ID = os.environ.get("MS_TENANT_ID", "common").strip()
MS_CLIENT_ID = os.environ.get("MS_CLIENT_ID", "").strip()
MS_CLIENT_SECRET = os.environ.get("MS_CLIENT_SECRET", "").strip()
MS_REDIRECT_URI = os.environ.get("MS_REDIRECT_URI", "").strip()
SSO_SCOPES = os.environ.get("SSO_SCOPES", "openid profile email User.Read").strip()
SSO_FETCH_ORG = os.environ.get("SSO_FETCH_ORG", "true").lower() in ("1", "true", "yes", "on")
SSO_PROMPT = os.environ.get("SSO_PROMPT", "select_account").strip()
ADMIN_EMAILS = {
    x.strip().lower()
    for x in os.environ.get("ADMIN_EMAILS", "").split(",")
    if x.strip()
}
OPERATOR_EMAILS = {
    x.strip().lower()
    for x in os.environ.get("OPERATOR_EMAILS", "john.kim@ezworks.co").split(",")
    if x.strip()
}
OPERATOR_DOMAINS = {
    x.strip().lower()
    for x in os.environ.get("OPERATOR_DOMAINS", "ezworks.co").split(",")
    if x.strip()
}
PROTECT_OPERATOR_ACCOUNTS = os.environ.get("PROTECT_OPERATOR_ACCOUNTS", "true").lower() in ("1", "true", "yes", "on")
EFFECTIVE_ADMIN_EMAILS = {*(ADMIN_EMAILS or set()), *(OPERATOR_EMAILS or set())}
LOCAL_LOGIN_ENABLED = os.environ.get("LOCAL_LOGIN_ENABLED", "true").lower() in ("1", "true", "yes", "on")
LOCAL_LOGIN_PASSWORD = os.environ.get("LOCAL_LOGIN_PASSWORD", ADMIN_PASSWORD).strip()
LOGIN_ROUTING_MODE = os.environ.get("LOGIN_ROUTING_MODE", "host").strip().lower()  # host | account
ACCOUNT_DOMAIN_REDIRECTS_RAW = os.environ.get("ACCOUNT_DOMAIN_REDIRECTS", "").strip()
DEFAULT_POST_LOGIN_REDIRECT = os.environ.get("DEFAULT_POST_LOGIN_REDIRECT", "/").strip() or "/"

init_user_db(str(USER_DB_PATH))
init_receipt_db(str(RECEIPT_DB_PATH))
for super_email in OPERATOR_EMAILS:
    upsert_user_by_email(
        db_path=str(USER_DB_PATH),
        email=super_email,
        name="Operator Admin",
        is_admin=True,
    )

oauth = OAuth(app) if OAuth else None
if oauth and MS_CLIENT_ID and MS_CLIENT_SECRET:
    oauth.register(
        name="microsoft",
        client_id=MS_CLIENT_ID,
        client_secret=MS_CLIENT_SECRET,
        server_metadata_url=f"https://login.microsoftonline.com/{MS_TENANT_ID}/v2.0/.well-known/openid-configuration",
        client_kwargs={"scope": SSO_SCOPES},
    )


def _is_sso_configured() -> bool:
    secret_invalid = _looks_like_secret_id(MS_CLIENT_SECRET)
    return bool(
        SSO_ENABLED
        and SSO_PROVIDER == "microsoft"
        and oauth
        and MS_CLIENT_ID
        and MS_CLIENT_SECRET
        and not secret_invalid
    )


def _looks_like_secret_id(value: str) -> bool:
    v = (value or "").strip()
    if len(v) != 36:
        return False
    parts = v.split("-")
    return [len(p) for p in parts] == [8, 4, 4, 4, 12]

# ─────────────────────────────────────────────
# Config helpers
# ─────────────────────────────────────────────
DEFAULT_CONFIG = {
    "company": {"name": "", "address": "", "phone": "", "logo_filename": ""},
    "fields": {
        "employee_name":  {"enabled": True,  "label": "Employee Name"},
        "department":     {"enabled": True,  "label": "Department"},
        "employee_id":    {"enabled": False, "label": "Employee ID"},
        "manager":        {"enabled": False, "label": "Manager / Supervisor"},
        "project":        {"enabled": False, "label": "Project / Cost Center"},
        "period":         {"enabled": True,  "label": "Expense Period"},
    },
    "expense_types": [
        {"id": "BREAKFAST",     "label": "Breakfast",         "enabled": True},
        {"id": "LUNCH",         "label": "Lunch",             "enabled": True},
        {"id": "DINNER",        "label": "Dinner",            "enabled": True},
        {"id": "ENTERTAINMENT", "label": "Entertainment",     "enabled": True},
        {"id": "LODGING",       "label": "Lodging / Hotel",   "enabled": True},
        {"id": "AIRFARE",       "label": "Airfare",           "enabled": True},
        {"id": "CAR_RENTAL",    "label": "Car Rental",        "enabled": True},
        {"id": "TAXI",          "label": "Taxi / Rideshare",  "enabled": True},
        {"id": "MILEAGE",       "label": "Mileage",           "enabled": True},
        {"id": "PARKING",       "label": "Parking / Tolls",   "enabled": True},
        {"id": "FUEL",          "label": "Fuel",              "enabled": False},
        {"id": "PHONE",         "label": "Phone / Internet",  "enabled": True},
        {"id": "OFFICE",        "label": "Office Supplies",   "enabled": False},
        {"id": "TIPS",          "label": "Tips / Gratuity",   "enabled": False},
        {"id": "MISCELLANEOUS", "label": "Miscellaneous",     "enabled": True},
    ],
    "modes": {
        "domestic":      {"enabled": True, "label": "Domestic (USD)"},
        "international": {"enabled": True, "label": "International"},
    },
    "login_page": {
        "interval_sec": 8,
        "slides": [],
    },
    "wizard": {
        "risk_priority": ["OCR_FAILED", "IMAGE_UNREADABLE", "LOW_CONFIDENCE", "NEEDS_REVIEW"],
        "cache_ttl_hours": 168,
    },
    "auth": {
        "allowed_email_domains": [],
        "admin_emails": [],
        "local_login_enabled": None,
        "local_login_password": "",
    },
}

DEFAULT_CATEGORY_ACCOUNT_CODES = {
    "BREAKFAST": "6100",
    "LUNCH": "6100",
    "DINNER": "6100",
    "ENTERTAINMENT": "6150",
    "LODGING": "6200",
    "AIRFARE": "6200",
    "CAR_RENTAL": "6205",
    "TAXI": "6205",
    "MILEAGE": "6210",
    "PARKING": "6210",
    "FUEL": "6300",
    "PHONE": "6400",
    "OFFICE": "6500",
    "TIPS": "6100",
    "MISCELLANEOUS": "6999",
}

ROLE_PERMISSIONS = {
    "general": {
        "receipt.create",
        "receipt.view_own",
        "report.submit",
        "report.view_own",
    },
    "accounting": {
        "receipt.view_all",
        "mapping.manage",
        "category.manage",
        "duplicate.resolve",
        "report.generate",
        "report.view_all",
    },
    "admin": {
        "company.manage",
        "template.manage",
        "user.manage",
        "policy.manage",
    },
}
ALL_PERMISSION_KEYS = sorted({p for perms in ROLE_PERMISSIONS.values() for p in perms})
DEFAULT_GENERAL_PERMISSIONS = set(ROLE_PERMISSIONS["general"])
DEFAULT_ADMIN_PERMISSIONS = {
    *ROLE_PERMISSIONS["general"],
    *ROLE_PERMISSIONS["accounting"],
    *ROLE_PERMISSIONS["admin"],
}


def _merge_with_default(data: dict) -> dict:
    merged = copy.deepcopy(DEFAULT_CONFIG)
    merged.update({k: v for k, v in data.items() if k in merged})
    for section in ("company", "fields", "modes", "login_page", "wizard", "auth"):
        if section in data and isinstance(data[section], dict):
            merged[section].update(data[section])
    if "expense_types" in data:
        merged["expense_types"] = data["expense_types"]
    return merged


def _get_request_host() -> str:
    forwarded = (request.headers.get("X-Forwarded-Host") or "").split(",")[0].strip()
    host = forwarded or request.host or "default"
    return host.split(":")[0].strip().lower()


def _host_to_config_key(host: str) -> str:
    host = (host or "default").lower()
    safe = "".join(ch if ch.isalnum() or ch in (".", "-", "_") else "_" for ch in host)
    return safe or "default"


def _config_path_for_host(host: str) -> Path:
    return CONFIG_DIR / f"{_host_to_config_key(host)}.json"


def _normalize_host_value(value: str) -> str:
    return (value or "").strip().lower().split(":")[0]


def _load_tenants_registry() -> list[dict]:
    """
    tenants.json format:
    [
      {"host":"lekpartners.com","name":"LEK Partners","redirect_url":"https://lekpartners.com","active":true}
    ]
    """
    if not TENANT_REGISTRY_PATH.exists():
        return []
    try:
        with open(TENANT_REGISTRY_PATH, encoding="utf-8") as f:
            raw = json.load(f)
        if not isinstance(raw, list):
            return []
        items = []
        seen = set()
        for row in raw:
            if not isinstance(row, dict):
                continue
            host = _normalize_host_value(str(row.get("host") or ""))
            if not host or host in seen:
                continue
            seen.add(host)
            items.append({
                "host": host,
                "name": str(row.get("name") or host).strip() or host,
                "redirect_url": str(row.get("redirect_url") or "").strip(),
                "active": bool(row.get("active", True)),
            })
        return items
    except Exception:
        return []


def _save_tenants_registry(items: list[dict]):
    cleaned = []
    seen = set()
    for row in (items or []):
        if not isinstance(row, dict):
            continue
        host = _normalize_host_value(str(row.get("host") or ""))
        if not host or host in seen:
            continue
        seen.add(host)
        cleaned.append({
            "host": host,
            "name": str(row.get("name") or host).strip() or host,
            "redirect_url": str(row.get("redirect_url") or "").strip(),
            "active": bool(row.get("active", True)),
        })
    with open(TENANT_REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(cleaned, f, ensure_ascii=False, indent=2)


def _ensure_tenant_known(host: str):
    host = _normalize_host_value(host)
    if not host:
        return
    rows = _load_tenants_registry()
    if any(x.get("host") == host for x in rows):
        return
    rows.append({"host": host, "name": host, "redirect_url": "", "active": True})
    _save_tenants_registry(rows)


def _redirect_url_from_registry(email_domain: str) -> str:
    """
    If tenant registry contains row with host matching email domain and redirect_url set,
    this can be used as account-based default redirect.
    """
    d = (email_domain or "").strip().lower()
    if not d:
        return ""
    for t in _load_tenants_registry():
        if not t.get("active", True):
            continue
        if t.get("host") == d and t.get("redirect_url"):
            return str(t.get("redirect_url"))
    return ""


def _parse_redirect_map(raw: str) -> dict:
    """
    ACCOUNT_DOMAIN_REDIRECTS format:
      'lekpartners.com=https://lek.example.com,kiotitractor.com=https://kioti.example.com'
    """
    mapping = {}
    for pair in (raw or "").split(","):
        item = pair.strip()
        if not item or "=" not in item:
            continue
        domain, target = item.split("=", 1)
        domain = domain.strip().lower()
        target = target.strip()
        if domain and target:
            mapping[domain] = target
    return mapping


ACCOUNT_DOMAIN_REDIRECTS = _parse_redirect_map(ACCOUNT_DOMAIN_REDIRECTS_RAW)


def load_config(host: str | None = None) -> dict:
    candidates = []
    if host:
        candidates.append(_config_path_for_host(host))
    candidates.append(CONFIG_PATH)
    for p in candidates:
        if not p.exists():
            continue
        try:
            with open(p, encoding="utf-8") as f:
                data = json.load(f)
            return _merge_with_default(data)
        except Exception:
            continue
    return copy.deepcopy(DEFAULT_CONFIG)


def save_config(cfg: dict, host: str | None = None):
    target = _config_path_for_host(host) if host else CONFIG_PATH
    with open(target, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def _email_domain(email: str) -> str:
    e = (email or "").strip().lower()
    if "@" not in e:
        return ""
    return e.split("@", 1)[1].strip().lower()


def _normalize_email_set(values) -> set:
    return {
        str(v).strip().lower()
        for v in (values or [])
        if str(v).strip() and "@" in str(v)
    }


def _normalize_domain_set(values) -> set:
    return {
        str(v).strip().lower()
        for v in (values or [])
        if str(v).strip()
    }


def _get_host_auth_settings(host: str) -> tuple[set, set]:
    cfg = load_config(host)
    auth = cfg.get("auth", {}) if isinstance(cfg.get("auth"), dict) else {}
    allowed_domains = _normalize_domain_set(auth.get("allowed_email_domains"))
    admin_emails = _normalize_email_set(auth.get("admin_emails"))
    return allowed_domains, admin_emails


def _host_local_login_policy(host: str) -> tuple[bool, str]:
    cfg = load_config(host)
    auth = cfg.get("auth", {}) if isinstance(cfg.get("auth"), dict) else {}
    enabled = auth.get("local_login_enabled")
    if enabled is None:
        enabled = LOCAL_LOGIN_ENABLED
    else:
        enabled = bool(enabled)
    password = str(auth.get("local_login_password") or "").strip() or LOCAL_LOGIN_PASSWORD
    return enabled, password


def _is_email_allowed_for_host(email: str, host: str) -> bool:
    if _is_operator_identity(email):
        return True
    allowed_domains, _ = _get_host_auth_settings(host)
    if not allowed_domains:
        return True
    return _email_domain(email) in allowed_domains


def _is_operator_session() -> bool:
    email = (session.get("user_email") or "").strip().lower()
    return _is_operator_identity(email)


def _is_admin_for_host(host: str) -> bool:
    user_id = session.get("user_id")
    if not user_id:
        return False
    user = get_user_by_id(str(USER_DB_PATH), int(user_id))
    if not user or not user.get("is_active"):
        _clear_session_user()
        return False
    email = (user.get("email") or "").strip().lower()
    if _is_operator_identity(email):
        return True
    _, tenant_admins = _get_host_auth_settings(host)
    return email in tenant_admins


def _resolve_target_host(req, body: dict | None = None) -> str:
    raw = ""
    if body and isinstance(body, dict):
        raw = str(body.get("target_host", "")).strip()
    if not raw:
        raw = str(req.args.get("host", "")).strip()
    return (raw or _get_request_host()).split(":")[0].strip().lower()


def _ensure_tenant_admin_allowed(target_host: str) -> tuple[bool, str]:
    current_host = _get_request_host()
    if target_host == current_host:
        return True, ""
    if _is_operator_session():
        return True, ""
    return False, "Cross-tenant admin management is allowed only for operator accounts."


def _set_tenant_admin_flag(target_host: str, email: str, enable: bool):
    _ensure_tenant_known(target_host)
    cfg = load_config(target_host)
    auth = cfg.setdefault("auth", {})
    admins = _normalize_email_set(auth.get("admin_emails"))
    if enable:
        admins.add(email)
    else:
        admins.discard(email)
    auth["admin_emails"] = sorted(admins)
    save_config(cfg, target_host)


def _get_effective_account_map(host: str) -> dict:
    effective = dict(DEFAULT_CATEGORY_ACCOUNT_CODES)
    for row in list_category_mappings(str(RECEIPT_DB_PATH), host):
        cat = str(row.get("category_id") or "").strip().upper()
        if not cat:
            continue
        code = str(row.get("account_code") or "").strip()
        if code:
            effective[cat] = code
    return effective


def _get_vendor_map(host: str) -> dict:
    rules = {}
    for row in list_vendor_mappings(str(RECEIPT_DB_PATH), host):
        key = str(row.get("vendor_name") or "").strip().upper()
        if not key:
            continue
        rules[key] = {
            "category": str(row.get("suggested_category") or "").strip().upper(),
            "account_code": str(row.get("account_code") or "").strip(),
        }
    return rules


def _current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_user_by_id(str(USER_DB_PATH), int(user_id))


def _effective_permissions_for_host(host: str) -> set:
    user = _current_user()
    if not user:
        return set()
    email = (user.get("email") or "").strip().lower()
    if _is_operator_identity(email):
        return set(ALL_PERMISSION_KEYS)

    explicit = get_user_host_permissions(str(USER_DB_PATH), int(user["id"]), host)
    if explicit:
        return explicit

    if _is_admin_for_host(host):
        return set(DEFAULT_ADMIN_PERMISSIONS)
    return set(DEFAULT_GENERAL_PERMISSIONS)


def _has_permission(host: str, permission: str) -> bool:
    if _is_operator_session():
        return True
    return permission in _effective_permissions_for_host(host)


def _require_any_permission(host: str, permissions: list[str]) -> tuple[bool, str]:
    if _is_operator_session():
        return True, ""
    perms = _effective_permissions_for_host(host)
    if any(p in perms for p in permissions):
        return True, ""
    return False, f"Permission denied. Required any of: {', '.join(permissions)}"


def _can_view_all_receipts(host: str) -> bool:
    return _is_operator_session() or _has_permission(host, "receipt.view_all")


def _can_view_own_receipts(host: str) -> bool:
    return _is_operator_session() or _has_permission(host, "receipt.view_own") or _has_permission(host, "receipt.view_all")


def _can_access_receipt_row(host: str, row: dict | None) -> bool:
    if not row:
        return False
    if _can_view_all_receipts(host):
        return True
    session_uid = int(session.get("user_id") or 0)
    return session_uid > 0 and int(row.get("uploader_user_id") or 0) == session_uid


def _guess_host_domains(host: str) -> set:
    h = (host or "").strip().lower()
    if not h or h in {"localhost", "127.0.0.1"}:
        return set()
    guesses = {h}
    if h.endswith(".local"):
        guesses.add(h[:-6] + ".com")
    return guesses


def _is_user_visible_for_host(email: str, host: str, allowed_domains: set, tenant_admins: set, operator_session: bool) -> bool:
    e = (email or "").strip().lower()
    if not e:
        return False
    if operator_session:
        return True
    session_email = (session.get("user_email") or "").strip().lower()
    if e == session_email:
        return True
    if _is_operator_identity(e):
        return False
    # Tenant admins are visible only when they are in the effective tenant scope.
    # This prevents localhost/internal admin pages from exposing customer users.
    if e in tenant_admins:
        session_domain = _email_domain(session_email)
        domain = _email_domain(e)
        if allowed_domains:
            return domain in allowed_domains
        guessed = _guess_host_domains(host)
        if guessed:
            return domain in guessed
        return bool(session_domain) and domain == session_domain
    domain = _email_domain(e)
    if allowed_domains:
        return domain in allowed_domains
    guessed = _guess_host_domains(host)
    if guessed:
        return domain in guessed
    # localhost/internal host fallback: scope to current session domain only.
    session_domain = _email_domain(session_email)
    return bool(session_domain) and domain == session_domain


def _resolve_post_login_redirect(email: str, requested_next: str) -> str:
    nxt = (requested_next or "").strip() or DEFAULT_POST_LOGIN_REDIRECT
    if LOGIN_ROUTING_MODE != "account":
        return nxt
    # Explicit deep-link(next) is respected; root-level landing is account-routed.
    if nxt not in ("/", ""):
        return nxt
    target = ACCOUNT_DOMAIN_REDIRECTS.get(_email_domain(email))
    if target:
        return target
    target = _redirect_url_from_registry(_email_domain(email))
    if target:
        return target
    return DEFAULT_POST_LOGIN_REDIRECT


def build_login_slides(cfg: dict) -> list:
    slides = []
    for item in cfg.get("login_page", {}).get("slides", []):
        if not isinstance(item, dict):
            continue
        fn = (item.get("filename") or "").strip()
        if not fn:
            continue
        p = LOGIN_BG_FOLDER / secure_filename(fn)
        if not p.exists():
            continue
        slides.append({
            "filename": p.name,
            "caption": str(item.get("caption") or ""),
            "url": f"/static/login_bg/{p.name}",
        })
    return slides


def _tenant_template_dir(host: str) -> Path:
    key = _host_to_config_key(host or "default")
    p = TENANT_TEMPLATE_ROOT / key
    p.mkdir(parents=True, exist_ok=True)
    return p


def _template_filename_for_mode(mode: str) -> str:
    mode_key = str(mode or "domestic").strip().lower()
    if mode_key == "international":
        return "template_international.xlsx"
    if mode_key == "domestic":
        return "template_domestic.xlsx"
    return "template.xlsx"


def get_template_path(mode: str = "domestic", host: str | None = None) -> Path | None:
    """Resolve template path in tenant scope first, then shared default fallback."""
    host_key = (host or _get_request_host()).strip().lower()
    mode_key = str(mode or "domestic").strip().lower()
    fname = _template_filename_for_mode(mode_key)
    candidates = [
        _tenant_template_dir(host_key) / fname,
        DEFAULT_TEMPLATE_DIR / fname,
    ]
    if mode_key in {"domestic", "international"}:
        candidates.extend(
            [
                _tenant_template_dir(host_key) / "template.xlsx",
                DEFAULT_TEMPLATE_DIR / "template.xlsx",
            ]
        )
    for p in candidates:
        if p.exists():
            return p
    return None


def allowed_file(filename, allowed_set):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed_set


def _extract_admin_password(req) -> str:
    body = req.get_json(silent=True) or {}
    return (body.get("password") or req.form.get("password", "")).strip()


def _is_admin_authenticated(host: str | None = None) -> bool:
    host = host or _get_request_host()
    return _is_admin_for_host(host)


def _is_user_authenticated() -> bool:
    user_id = session.get("user_id")
    if user_id:
        user = get_user_by_id(str(USER_DB_PATH), int(user_id))
        if user and user.get("is_active"):
            return True
        email = (session.get("user_email") or "").strip().lower()
        if email:
            rebuilt = upsert_user_by_email(
                db_path=str(USER_DB_PATH),
                email=email,
                name=(session.get("user_name") or "").strip(),
                is_admin=bool(session.get("is_admin")),
            )
            if rebuilt and rebuilt.get("is_active"):
                _set_session_user(rebuilt, provider=str(session.get("auth_provider") or "local"))
                return True
        _clear_session_user()
        return False
    email_only = (session.get("user_email") or "").strip().lower()
    if email_only:
        rebuilt = upsert_user_by_email(
            db_path=str(USER_DB_PATH),
            email=email_only,
            name=(session.get("user_name") or "").strip(),
            is_admin=bool(session.get("is_admin")),
        )
        if rebuilt and rebuilt.get("is_active"):
            _set_session_user(rebuilt, provider=str(session.get("auth_provider") or "local"))
            return True
        _clear_session_user()
        return False
    if session.get("is_admin"):
        _clear_session_user()
    return False


def _set_session_user(user: dict, provider: str = "local"):
    session.permanent = True
    session["user_id"] = user.get("id")
    session["user_email"] = user.get("email")
    session["user_name"] = user.get("name") or ""
    session["auth_provider"] = provider
    session["is_admin"] = bool(user.get("is_admin"))


def _clear_session_user():
    for key in ("user_id", "user_email", "user_name", "auth_provider", "is_admin"):
        session.pop(key, None)


def _build_abs_url(path: str) -> str:
    base = request.url_root.rstrip("/")
    if not path.startswith("/"):
        path = "/" + path
    return f"{base}{path}"


def _effective_ms_redirect_uri() -> str:
    """
    Use host-matching callback URI to avoid cross-host session-cookie loss.
    If configured MS_REDIRECT_URI host differs from current request host,
    fallback to current host callback URL.
    """
    dynamic_uri = _build_abs_url("/auth/callback/microsoft")
    if not MS_REDIRECT_URI:
        return dynamic_uri
    try:
        cfg = urlparse.urlparse(MS_REDIRECT_URI)
        req = urlparse.urlparse(dynamic_uri)
        cfg_host = (cfg.hostname or "").strip().lower()
        req_host = (req.hostname or "").strip().lower()
        if cfg_host and req_host and cfg_host != req_host:
            return dynamic_uri
    except Exception:
        return dynamic_uri
    return MS_REDIRECT_URI


def _is_operator_identity(email: str) -> bool:
    e = (email or "").strip().lower()
    if not e:
        return False
    if e in OPERATOR_EMAILS:
        return True
    if "@" in e and e.split("@", 1)[1] in OPERATOR_DOMAINS:
        return True
    return False


def _fetch_ms_org_name(access_token: str) -> str:
    if not access_token or not SSO_FETCH_ORG:
        return ""
    try:
        req = urlrequest.Request(
            "https://graph.microsoft.com/v1.0/organization?$select=id,displayName",
            headers={"Authorization": f"Bearer {access_token}"},
        )
        with urlrequest.urlopen(req, timeout=5) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
        values = payload.get("value") or []
        if values and values[0].get("displayName"):
            return str(values[0]["displayName"])
    except Exception:
        pass
    return ""


def _exchange_ms_code_directly(code: str, redirect_uri: str) -> dict:
    token_url = f"https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token"
    form = {
        "grant_type": "authorization_code",
        "client_id": MS_CLIENT_ID,
        "client_secret": MS_CLIENT_SECRET,
        "code": code,
        "redirect_uri": redirect_uri,
        "scope": SSO_SCOPES,
    }
    data = urlparse.urlencode(form).encode("utf-8")
    req = urlrequest.Request(
        token_url,
        data=data,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
    )
    try:
        with urlrequest.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urlerror.HTTPError as e:
        detail = ""
        try:
            detail = e.read().decode("utf-8")
        except Exception:
            detail = str(e)
        raise ValueError(f"{e.code} {e.reason}: {detail}")


def _fetch_ms_me(access_token: str) -> dict:
    req = urlrequest.Request(
        "https://graph.microsoft.com/v1.0/me",
        headers={"Authorization": f"Bearer {access_token}"},
    )
    with urlrequest.urlopen(req, timeout=10) as resp:
        return json.loads(resp.read().decode("utf-8"))


@app.before_request
def require_login_for_api():
    if request.path.startswith("/api/") and not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401


# ─────────────────────────────────────────────
# Main App Routes
# ─────────────────────────────────────────────

@app.route("/")
def index():
    if not _is_user_authenticated():
        return redirect("/login?next=/")
    return redirect("/dashboard")


@app.route("/dashboard")
def dashboard_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/dashboard")
    return _render_inbox_page("dashboard.html", active_module="dashboard")


@app.route("/legacy")
def legacy_index():
    if not _is_user_authenticated():
        return redirect("/login?next=/legacy")
    # Legacy single-page flow is deprecated; keep module-based UX only.
    return redirect("/dashboard")


@app.route("/inbox")
def inbox_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/inbox")
    return redirect("/inbox/review")


def _render_inbox_page(template_name: str, active_module: str = "inbox"):
    cfg = load_config(_get_request_host())
    host = _get_request_host()
    logo_url = (
        f"/static/logos/{cfg['company']['logo_filename']}"
        if cfg["company"].get("logo_filename") else None
    )
    can_access_admin_module, _ = _require_any_permission(
        host,
        ["mapping.manage", "category.manage", "company.manage", "template.manage", "user.manage", "policy.manage"],
    )
    return render_template(
        template_name,
        config=cfg,
        logo_url=logo_url,
        active_module=active_module,
        can_access_admin_module=bool(can_access_admin_module),
    )


@app.route("/inbox/upload")
def inbox_upload_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/inbox/upload")
    return redirect("/reports/wizard/upload")


@app.route("/inbox/review")
def inbox_review_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/inbox/review")
    return _render_inbox_page("inbox_review.html", active_module="inbox")


@app.route("/reports")
def reports_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports")
    return _render_inbox_page("inbox_reports.html", active_module="reports")


@app.route("/reports/<int:report_id>")
def report_detail_page(report_id: int):
    if not _is_user_authenticated():
        return redirect(f"/login?next=/reports/{report_id}")
    return _render_inbox_page("report_detail.html", active_module="reports")


@app.route("/inbox/reports")
def inbox_reports_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/inbox/reports")
    return redirect("/reports")


@app.route("/inbox/settings")
def inbox_settings_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/inbox/settings")
    host = _get_request_host()
    ok, _ = _require_any_permission(host, ["mapping.manage", "category.manage", "company.manage", "template.manage", "user.manage", "policy.manage"])
    if not ok:
        return redirect("/inbox/review")
    return _render_inbox_page("inbox_settings.html", active_module="admin")


@app.route("/reports/wizard")
def reports_wizard_entry():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard")
    return redirect("/reports/wizard/upload")


@app.route("/reports/wizard/settings")
def reports_wizard_settings_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard/settings")
    return redirect("/reports/wizard/upload")


@app.route("/reports/wizard/upload")
def reports_wizard_upload_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard/upload")
    return _render_inbox_page("report_wizard_upload.html", active_module="reports")


@app.route("/reports/wizard/ocr")
def reports_wizard_ocr_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard/ocr")
    return redirect("/reports/wizard/upload")


@app.route("/reports/wizard/review")
def reports_wizard_review_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard/review")
    return _render_inbox_page("report_wizard_review.html", active_module="reports")


@app.route("/reports/wizard/generate")
def reports_wizard_generate_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/reports/wizard/generate")
    return _render_inbox_page("report_wizard_generate.html", active_module="reports")


@app.route("/login")
def login_page():
    if _is_user_authenticated():
        return redirect(request.args.get("next", "/"))
    host = _get_request_host()
    local_enabled, _ = _host_local_login_policy(host)
    cfg = load_config(_get_request_host())
    interval_sec = int(cfg.get("login_page", {}).get("interval_sec") or 8)
    interval_sec = min(max(interval_sec, 3), 120)
    return render_template(
        "login.html",
        slides=build_login_slides(cfg),
        interval_sec=interval_sec,
        sso_enabled=SSO_ENABLED and SSO_PROVIDER == "microsoft",
        sso_configured=_is_sso_configured(),
        sso_missing=(
            [x for x in ("MS_CLIENT_ID", "MS_CLIENT_SECRET") if not {"MS_CLIENT_ID": MS_CLIENT_ID, "MS_CLIENT_SECRET": MS_CLIENT_SECRET}[x]]
            + (["MS_CLIENT_SECRET(Value)"] if _looks_like_secret_id(MS_CLIENT_SECRET) else [])
        ),
        sso_callback=_effective_ms_redirect_uri(),
        local_login_enabled=local_enabled,
    )


@app.route("/api/config")
def api_config():
    """프론트엔드가 동적으로 사용하는 설정 반환"""
    cfg = load_config(_get_request_host())
    logo_url = (
        f"/static/logos/{cfg['company']['logo_filename']}"
        if cfg["company"].get("logo_filename") else None
    )
    return jsonify({**cfg, "logo_url": logo_url})


@app.route("/api/status")
def api_status():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    host = _get_request_host()
    return jsonify({
        "api_key_set":       bool(api_key),
        "api_key_preview":   (api_key[:12] + "...") if api_key else None,
        "template_domestic": get_template_path("domestic", host=host)      is not None,
        "template_intl":     get_template_path("international", host=host) is not None,
    })


# ─── Receipt Inbox API (Tenant-scoped) ───

def _reconcile_duplicate_receipt(host: str, receipt_id: int):
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return None
    dup = find_duplicate_receipt(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        receipt_id=receipt_id,
        merchant=row.get("merchant"),
        amount=int(row.get("amount") or 0),
        date=row.get("date"),
    )
    if dup:
        return update_receipt(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            receipt_id=receipt_id,
            fields={
                "status": "duplicate",
                "duplicate_of": int(dup["id"]),
                "lifecycle_state": "NEEDS_REVIEW",
            },
        )
    if row.get("status") == "duplicate":
        fallback_status = "processed" if (row.get("date") and int(row.get("amount") or 0) > 0) else "needs_review"
        return update_receipt(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            receipt_id=receipt_id,
            fields={
                "status": fallback_status,
                "duplicate_of": None,
                "lifecycle_state": _derive_lifecycle_state(row, {"status": fallback_status, "duplicate_of": None}),
            },
        )
    return row


def _derive_lifecycle_state(row: dict | None, pending_fields: dict | None = None) -> str:
    data = dict(row or {})
    for k, v in (pending_fields or {}).items():
        data[k] = v
    report_status = str(data.get("report_status") or "unassigned").strip().lower()
    if report_status == "assigned":
        return "ASSIGNED_TO_REPORT"
    status = str(data.get("status") or "").strip().lower()
    confidence = str(data.get("confidence") or "").strip().lower()
    ocr_error = str(data.get("ocr_error") or "").strip()
    if status == "ocr_processing":
        return "OCR_PROCESSING"
    if ocr_error:
        if "low" in ocr_error.lower() or confidence == "low":
            return "LOW_CONFIDENCE"
        if "unreadable" in ocr_error.lower():
            return "IMAGE_UNREADABLE"
        return "OCR_FAILED"
    if status == "duplicate":
        return "NEEDS_REVIEW"
    if status == "processed":
        return "READY_FOR_REPORT"
    if status == "needs_review":
        return "NEEDS_REVIEW"
    return "UPLOADED"


def _within_last_days(iso_text: str, days: int | None) -> bool:
    if not days or int(days) <= 0:
        return True
    raw = str(iso_text or "").strip()
    if not raw:
        return False
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
    except Exception:
        return False
    cutoff = datetime.now(timezone.utc) - timedelta(days=int(days))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt >= cutoff

@app.route("/api/inbox/upload", methods=["POST"])
def inbox_upload_receipts():
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["receipt.create"])
    if not ok:
        return jsonify({"error": err}), 403
    if "receipts" not in request.files:
        return jsonify({"error": "No receipt files provided."}), 400
    host_key = _host_to_config_key(host)
    host_dir = INBOX_UPLOAD_FOLDER / host_key
    host_dir.mkdir(parents=True, exist_ok=True)

    uploader_user_id = session.get("user_id")
    uploader_email = (session.get("user_email") or "").strip().lower()
    rows = []

    for file in request.files.getlist("receipts"):
        if not file or not allowed_file(file.filename, ALLOWED_IMAGE_EXT):
            continue
        orig_name = secure_filename(file.filename)
        file_bytes = file.read()
        file.stream.seek(0)
        file_hash = hashlib.sha256(file_bytes).hexdigest() if file_bytes else ""
        dup = find_receipt_by_hash(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            file_hash=file_hash,
            uploader_user_id=int(uploader_user_id) if uploader_user_id else None,
            only_unassigned=True,
        )
        if dup:
            rows.append(dup)
            continue
        stored_name = f"{uuid.uuid4().hex[:8]}_{orig_name}"
        save_path = host_dir / stored_name
        file.save(str(save_path))
        row = create_receipt(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            uploader_user_id=int(uploader_user_id) if uploader_user_id else None,
            uploader_email=uploader_email,
            orig_filename=orig_name,
            stored_filename=stored_name,
            file_path=str(save_path),
            file_hash=file_hash,
        )
        row = update_receipt(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            receipt_id=int(row["id"]),
            fields={"lifecycle_state": "UPLOADED"},
        ) or row
        rows.append(row)

    if not rows:
        return jsonify({"error": "No valid image files found."}), 400
    return jsonify({"status": "ok", "count": len(rows), "receipts": rows})


@app.route("/api/inbox/receipts", methods=["GET"])
def inbox_list_receipts():
    host = _get_request_host()
    can_view_own = _can_view_own_receipts(host)
    can_view_all = _can_view_all_receipts(host)
    if not can_view_own:
        return jsonify({"error": "Permission denied. Missing receipt view permission."}), 403
    status = (request.args.get("status") or "").strip() or None
    lifecycle_state = (request.args.get("lifecycle_state") or "").strip().upper() or None
    report_status = (request.args.get("report_status") or "").strip() or None
    date_from = (request.args.get("date_from") or "").strip() or None
    date_to = (request.args.get("date_to") or "").strip() or None
    category = (request.args.get("category") or "").strip() or None
    merchant = (request.args.get("merchant") or "").strip() or None
    min_amount_raw = (request.args.get("min_amount") or "").strip()
    max_amount_raw = (request.args.get("max_amount") or "").strip()
    user_email = (request.args.get("user_email") or "").strip().lower() or None
    try:
        min_amount = int(min_amount_raw) if min_amount_raw else None
    except Exception:
        return jsonify({"error": "min_amount must be an integer"}), 400
    try:
        max_amount = int(max_amount_raw) if max_amount_raw else None
    except Exception:
        return jsonify({"error": "max_amount must be an integer"}), 400
    q = (request.args.get("q") or "").strip() or None
    limit = int(request.args.get("limit") or 200)
    items = list_receipts(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        status=status,
        lifecycle_state=lifecycle_state,
        report_status=report_status,
        date_from=date_from,
        date_to=date_to,
        category=category,
        merchant=merchant,
        min_amount=min_amount,
        max_amount=max_amount,
        search=q,
        limit=limit,
    )
    account_map = _get_effective_account_map(host)
    for item in items:
        item["account_code"] = account_map.get((item.get("category") or "").upper(), "")
    if can_view_all:
        if user_email:
            items = [x for x in items if (x.get("uploader_email") or "").strip().lower() == user_email]
    else:
        uid = session.get("user_id")
        items = [x for x in items if int(x.get("uploader_user_id") or 0) == int(uid or 0)]
    return jsonify({"receipts": items, "count": len(items), "host": host})


@app.route("/api/inbox/category-mappings", methods=["GET"])
def inbox_list_category_mappings():
    host = _get_request_host()
    perms = _effective_permissions_for_host(host)
    cfg = load_config(host)
    rows = list_category_mappings(str(RECEIPT_DB_PATH), host)
    by_cat = {r["category_id"]: r for r in rows}
    merged = []
    for t in cfg.get("expense_types", []):
        if not t.get("id"):
            continue
        cat = str(t["id"]).upper()
        row = by_cat.get(cat, {})
        default_code = DEFAULT_CATEGORY_ACCOUNT_CODES.get(cat, "")
        merged.append(
            {
                "category_id": cat,
                "category_label": str(t.get("label") or cat),
                "default_account_code": default_code,
                "account_code": row.get("account_code", "") or default_code,
                "updated_at": row.get("updated_at"),
                "overridden": bool(row.get("account_code")),
            }
        )
    return jsonify({"host": host, "mappings": merged, "can_manage": ("mapping.manage" in perms or "category.manage" in perms)})


@app.route("/api/inbox/category-mappings", methods=["POST"])
def inbox_upsert_category_mappings():
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["mapping.manage", "category.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    body = request.get_json(silent=True) or {}
    mappings = body.get("mappings")
    if not isinstance(mappings, list) or not mappings:
        return jsonify({"error": "mappings list is required"}), 400
    updated = 0
    for item in mappings:
        if not isinstance(item, dict):
            continue
        cat = str(item.get("category_id") or "").strip().upper()
        label = str(item.get("category_label") or cat).strip()
        code = str(item.get("account_code") or "").strip()
        if not cat:
            continue
        upsert_category_mapping(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            category_id=cat,
            category_label=label,
            account_code=code,
        )
        updated += 1
    return jsonify({"status": "ok", "updated": updated})


@app.route("/api/inbox/vendor-mappings", methods=["GET"])
def inbox_list_vendor_mappings():
    host = _get_request_host()
    rows = list_vendor_mappings(str(RECEIPT_DB_PATH), host)
    perms = _effective_permissions_for_host(host)
    return jsonify({"host": host, "mappings": rows, "can_manage": ("mapping.manage" in perms or "category.manage" in perms)})


@app.route("/api/inbox/vendor-mappings", methods=["POST"])
def inbox_upsert_vendor_mappings():
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["mapping.manage", "category.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    body = request.get_json(silent=True) or {}
    mappings = body.get("mappings")
    if not isinstance(mappings, list):
        return jsonify({"error": "mappings(list) is required"}), 400
    updated = 0
    for item in mappings:
        if not isinstance(item, dict):
            continue
        vendor = str(item.get("vendor_name") or "").strip().upper()
        if not vendor:
            continue
        upsert_vendor_mapping(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            vendor_name=vendor,
            suggested_category=str(item.get("suggested_category") or "").strip().upper(),
            account_code=str(item.get("account_code") or "").strip(),
        )
        updated += 1
    return jsonify({"status": "ok", "updated": updated})


@app.route("/api/inbox/permissions", methods=["GET"])
def inbox_permissions():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    perms = sorted(_effective_permissions_for_host(host))
    return jsonify(
        {
            "host": host,
            "operator": _is_operator_session(),
            "permissions": perms,
            "can_manage_mappings": ("mapping.manage" in perms or "category.manage" in perms),
        }
    )


@app.route("/api/inbox/filter-presets", methods=["GET"])
def inbox_list_filter_presets():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    personal = list_inbox_filter_presets(str(USER_DB_PATH), uid, host)
    shared = list_shared_inbox_filter_presets(str(USER_DB_PATH), host)
    can_admin, _ = _require_any_permission(host, ["user.manage", "policy.manage"])
    user_email = (session.get("user_email") or "").strip().lower()
    shared_view = []
    for s in shared:
        is_owner = int(s.get("created_by_user_id") or 0) == uid or (s.get("created_by_email") or "").strip().lower() == user_email
        is_locked = bool(s.get("is_locked"))
        can_edit = bool(can_admin or (is_owner and not is_locked))
        can_delete = bool(can_admin or (is_owner and not is_locked))
        shared_view.append({**s, "is_owner": is_owner, "can_edit": can_edit, "can_delete": can_delete})
    combined = (
        [{"scope": "personal", **x} for x in personal]
        + [{"scope": "shared", **x} for x in shared_view]
    )
    return jsonify({"host": host, "presets": combined, "personal": personal, "shared": shared_view, "count": len(combined)})


@app.route("/api/inbox/filter-presets", methods=["POST"])
def inbox_save_filter_preset():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    body = request.get_json(silent=True) or {}
    name = str(body.get("preset_name") or "").strip()
    filters = body.get("filters") if isinstance(body.get("filters"), dict) else {}
    scope = str(body.get("scope") or "personal").strip().lower()
    lock_requested = body.get("is_locked")
    if not name:
        return jsonify({"error": "preset_name is required"}), 400
    if scope == "shared":
        existing = get_shared_inbox_filter_preset(str(USER_DB_PATH), host, name)
        is_owner = bool(existing and int(existing.get("created_by_user_id") or 0) == uid)
        ok_admin, _ = _require_any_permission(host, ["user.manage", "policy.manage"])
        existing_locked = bool(existing and existing.get("is_locked"))
        if existing and existing_locked and not ok_admin:
            return jsonify({"error": "Locked shared preset can be changed only by admin."}), 403
        if existing:
            if not (is_owner or ok_admin):
                return jsonify({"error": "Only preset owner or admin can update shared preset."}), 403
        else:
            if not ok_admin:
                return jsonify({"error": "Admin permission required to create shared preset."}), 403
        if lock_requested is not None and not ok_admin:
            return jsonify({"error": "Only admin can lock/unlock shared preset."}), 403
        row = upsert_shared_inbox_filter_preset(
            str(USER_DB_PATH),
            host,
            name,
            filters,
            created_by_user_id=uid,
            created_by_email=(session.get("user_email") or ""),
            is_locked=bool(lock_requested) if lock_requested is not None else None,
        )
        try:
            action = "create" if not existing else ("lock_toggle" if bool(existing.get("is_locked")) != bool(row.get("is_locked")) else "update")
            create_shared_inbox_filter_preset_audit_log(
                str(USER_DB_PATH),
                host,
                name,
                actor_user_id=uid,
                actor_email=(session.get("user_email") or ""),
                action=action,
                before_json=json.dumps(existing or {}, ensure_ascii=False),
                after_json=json.dumps(row or {}, ensure_ascii=False),
            )
        except Exception:
            pass
    else:
        row = upsert_inbox_filter_preset(str(USER_DB_PATH), uid, host, name, filters)
        scope = "personal"
    return jsonify({"status": "ok", "host": host, "scope": scope, "preset": row})


@app.route("/api/inbox/filter-presets/<preset_name>", methods=["DELETE"])
def inbox_delete_filter_preset(preset_name: str):
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    scope = str(request.args.get("scope") or "personal").strip().lower()
    if scope == "shared":
        existing = get_shared_inbox_filter_preset(str(USER_DB_PATH), host, preset_name)
        if not existing:
            return jsonify({"status": "ok", "scope": "shared", "deleted": False})
        is_owner = int(existing.get("created_by_user_id") or 0) == uid
        ok_admin, _ = _require_any_permission(host, ["user.manage", "policy.manage"])
        if bool(existing.get("is_locked")) and not ok_admin:
            return jsonify({"error": "Locked shared preset can be deleted only by admin."}), 403
        if not (is_owner or ok_admin):
            return jsonify({"error": "Only preset owner or admin can delete shared preset."}), 403
        try:
            create_shared_inbox_filter_preset_audit_log(
                str(USER_DB_PATH),
                host,
                preset_name,
                actor_user_id=uid,
                actor_email=(session.get("user_email") or ""),
                action="delete",
                before_json=json.dumps(existing or {}, ensure_ascii=False),
                after_json="{}",
            )
        except Exception:
            pass
        ok = delete_shared_inbox_filter_preset(str(USER_DB_PATH), host, preset_name)
    else:
        ok = delete_inbox_filter_preset(str(USER_DB_PATH), uid, host, preset_name)
        scope = "personal"
    return jsonify({"status": "ok", "scope": scope, "deleted": bool(ok)})


@app.route("/api/inbox/filter-presets/<preset_name>/lock", methods=["PATCH"])
def inbox_toggle_shared_filter_preset_lock(preset_name: str):
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    ok_admin, err = _require_any_permission(host, ["user.manage", "policy.manage"])
    if not ok_admin:
        return jsonify({"error": err}), 403
    body = request.get_json(silent=True) or {}
    existing = get_shared_inbox_filter_preset(str(USER_DB_PATH), host, preset_name)
    if not existing:
        return jsonify({"error": "Shared preset not found."}), 404
    locked = bool(body.get("is_locked"))
    row = upsert_shared_inbox_filter_preset(
        str(USER_DB_PATH),
        host,
        preset_name,
        existing.get("filters") or {},
        created_by_user_id=existing.get("created_by_user_id"),
        created_by_email=existing.get("created_by_email") or "",
        is_locked=locked,
    )
    try:
        create_shared_inbox_filter_preset_audit_log(
            str(USER_DB_PATH),
            host,
            preset_name,
            actor_user_id=uid,
            actor_email=(session.get("user_email") or ""),
            action="lock_toggle",
            before_json=json.dumps(existing or {}, ensure_ascii=False),
            after_json=json.dumps(row or {}, ensure_ascii=False),
        )
    except Exception:
        pass
    return jsonify({"status": "ok", "preset": row})


@app.route("/api/inbox/wizard-preferences", methods=["GET"])
def inbox_get_wizard_preferences():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    row = get_user_wizard_preference(str(USER_DB_PATH), uid, host)
    return jsonify({"host": host, "preference": row})


@app.route("/api/inbox/wizard-preferences", methods=["POST"])
def inbox_save_wizard_preferences():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    host = _get_request_host()
    body = request.get_json(silent=True) or {}
    mode = str(body.get("mode") or "default").strip().lower()
    rp = body.get("risk_priority") if isinstance(body.get("risk_priority"), list) else []
    allowed = {"OCR_FAILED", "IMAGE_UNREADABLE", "LOW_CONFIDENCE", "NEEDS_REVIEW"}
    risk_priority = [str(x).strip().upper() for x in rp if str(x).strip().upper() in allowed]
    row = upsert_user_wizard_preference(str(USER_DB_PATH), uid, host, mode, risk_priority)
    return jsonify({"status": "ok", "host": host, "preference": row})


@app.route("/api/inbox/receipts/<int:receipt_id>", methods=["GET"])
def inbox_get_receipt(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt view permission."}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot view receipt uploaded by another user."}), 403
    return jsonify({"host": host, "receipt": row})


@app.route("/api/inbox/receipts/<int:receipt_id>", methods=["PATCH"])
def inbox_update_receipt(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt edit permission."}), 403
    body = request.get_json(silent=True) or {}
    allowed_status = {"processed", "needs_review", "duplicate"}
    allowed_report_status = {"assigned", "unassigned"}
    fields = {}
    for key in ("date", "merchant", "currency", "category", "memo", "confidence", "ocr_error"):
        if key in body:
            fields[key] = body.get(key)
    if "amount" in body:
        try:
            fields["amount"] = max(0, int(body.get("amount") or 0))
        except Exception:
            return jsonify({"error": "amount must be an integer"}), 400
    if "status" in body:
        status = str(body.get("status") or "").strip().lower()
        if status not in allowed_status:
            return jsonify({"error": "invalid status"}), 400
        fields["status"] = status
    if "report_status" in body:
        report_state = str(body.get("report_status") or "").strip().lower()
        if report_state not in allowed_report_status:
            return jsonify({"error": "invalid report_status"}), 400
        fields["report_status"] = report_state

    current = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not current:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, current):
        return jsonify({"error": "Cannot edit receipt uploaded by another user."}), 403

    if "status" in fields and fields["status"] == "duplicate":
        ok, err = _require_any_permission(host, ["duplicate.resolve"])
        if not ok:
            return jsonify({"error": err}), 403

    fields["lifecycle_state"] = _derive_lifecycle_state(current, fields)
    item = update_receipt(str(RECEIPT_DB_PATH), host, receipt_id, fields)
    if not item:
        return jsonify({"error": "Receipt not found."}), 404
    if {"date", "merchant", "amount"} & set(fields.keys()):
        item = _reconcile_duplicate_receipt(host, receipt_id)
    changed_fields = []
    for k in fields.keys():
        if current.get(k) != item.get(k):
            changed_fields.append(k)
    if changed_fields:
        create_receipt_audit_log(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            receipt_id=receipt_id,
            actor_user_id=int(session.get("user_id")) if session.get("user_id") else None,
            actor_email=(session.get("user_email") or ""),
            action="update",
            changed_fields=changed_fields,
            before_json=json.dumps({k: current.get(k) for k in changed_fields}, ensure_ascii=False),
            after_json=json.dumps({k: item.get(k) for k in changed_fields}, ensure_ascii=False),
        )
    return jsonify({"status": "ok", "receipt": item})


@app.route("/api/inbox/receipts/<int:receipt_id>", methods=["DELETE"])
def inbox_delete_receipt(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt delete permission."}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot delete receipt uploaded by another user."}), 403
    if row.get("report_status") == "assigned":
        return jsonify({"error": "Assigned receipts cannot be deleted. Unassign first."}), 400

    image_path = Path(row.get("file_path") or "")
    deleted = delete_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not deleted:
        return jsonify({"error": "Receipt not found."}), 404

    try:
        if image_path.exists() and image_path.is_file():
            image_path.unlink()
    except Exception:
        pass

    return jsonify({"status": "ok", "deleted_id": int(receipt_id)})


@app.route("/api/inbox/receipts/<int:receipt_id>/audit", methods=["GET"])
def inbox_receipt_audit_logs(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt view permission."}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot view audit of another user's receipt."}), 403
    logs = list_receipt_audit_logs(str(RECEIPT_DB_PATH), host, receipt_id, limit=int(request.args.get("limit") or 50))
    return jsonify({"host": host, "receipt_id": receipt_id, "logs": logs, "count": len(logs)})


@app.route("/api/inbox/ocr/<int:receipt_id>", methods=["POST"])
def inbox_ocr_receipt(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt OCR permission."}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot OCR receipt uploaded by another user."}), 403
    body = request.get_json(silent=True) or {}
    force = str(request.args.get("force") or body.get("force") or "").strip().lower() in ("1", "true", "yes", "on")
    lc = str(row.get("lifecycle_state") or "").upper()
    status_now = str(row.get("status") or "").strip().lower()
    confidence_now = str(row.get("confidence") or "").strip().lower()
    if not force:
        if lc in {"READY_FOR_REPORT", "ASSIGNED_TO_REPORT"}:
            return jsonify({"error": "Receipt already processed. OCR rerun is blocked to avoid duplicate cost."}), 400
        if status_now == "processed" and confidence_now in {"high", "medium", "manual"} and not str(row.get("ocr_error") or "").strip():
            return jsonify({"error": "Receipt already has successful OCR result. Use force=true to rerun."}), 400
    image_path = Path(row.get("file_path") or "")
    if not image_path.exists():
        return jsonify({"error": "Receipt image file not found."}), 404
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "ANTHROPIC_API_KEY is missing."}), 500
    update_receipt(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        receipt_id=receipt_id,
        fields={"status": "ocr_processing", "lifecycle_state": "OCR_PROCESSING"},
    )

    result = process_receipt_image(str(image_path), api_key)
    merchant_raw = str(result.get("merchant") or "").strip().upper()
    if merchant_raw:
        vendor_map = _get_vendor_map(host)
        vendor_hit = vendor_map.get(merchant_raw)
        if not vendor_hit:
            for vkey, vrule in vendor_map.items():
                if vkey and vkey in merchant_raw:
                    vendor_hit = vrule
                    break
        if vendor_hit:
            if vendor_hit.get("category"):
                result["category"] = vendor_hit["category"]
            acct = (vendor_hit.get("account_code") or "").strip()
            if acct:
                memo_text = str(result.get("memo") or "")
                if f"AC:{acct}" not in memo_text:
                    result["memo"] = (f"{memo_text} | AC:{acct}").strip(" |")
    has_error = bool(result.get("error"))
    date_val = result.get("date")
    amount_val = int(result.get("amount") or 0)
    next_status = "processed" if (not has_error and date_val and amount_val > 0) else "needs_review"
    updated = update_receipt(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        receipt_id=receipt_id,
        fields={
            "date": date_val,
            "merchant": (result.get("merchant") or ""),
            "amount": amount_val,
            "currency": (result.get("currency") or "USD"),
            "category": (result.get("category") or "MISCELLANEOUS"),
            "memo": (result.get("memo") or ""),
            "confidence": (result.get("confidence") or "low"),
            "status": next_status,
            "ocr_error": (result.get("error") or ""),
            "lifecycle_state": _derive_lifecycle_state(
                row,
                {
                    "status": next_status,
                    "ocr_error": (result.get("error") or ""),
                    "confidence": (result.get("confidence") or "low"),
                },
            ),
        },
    )
    updated = _reconcile_duplicate_receipt(host, receipt_id)
    return jsonify({"status": "ok", "receipt": updated, "ocr": result})


@app.route("/api/inbox/receipts/<int:receipt_id>/ignore-duplicate", methods=["POST"])
def inbox_ignore_duplicate(receipt_id: int):
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["duplicate.resolve"])
    if not ok:
        return jsonify({"error": err}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot modify receipt uploaded by another user."}), 403
    next_status = "processed" if (row.get("date") and int(row.get("amount") or 0) > 0) else "needs_review"
    updated = update_receipt(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        receipt_id=receipt_id,
        fields={
            "status": next_status,
            "duplicate_of": None,
            "lifecycle_state": _derive_lifecycle_state(row, {"status": next_status, "duplicate_of": None}),
        },
    )
    return jsonify({"status": "ok", "receipt": updated})


@app.route("/api/inbox/image/<int:receipt_id>")
def inbox_receipt_image(receipt_id: int):
    host = _get_request_host()
    if not _can_view_own_receipts(host):
        return jsonify({"error": "Permission denied. Missing receipt view permission."}), 403
    row = get_receipt(str(RECEIPT_DB_PATH), host, receipt_id)
    if not row:
        return jsonify({"error": "Receipt not found."}), 404
    if not _can_access_receipt_row(host, row):
        return jsonify({"error": "Cannot view receipt uploaded by another user."}), 403
    image_path = Path(row.get("file_path") or "")
    if not image_path.exists():
        return jsonify({"error": "Image not found."}), 404
    ext = image_path.suffix.lower()
    mime_map = {
        ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
        ".webp": "image/webp", ".heic": "image/heic", ".gif": "image/gif",
    }
    return send_file(str(image_path), mimetype=mime_map.get(ext, "image/jpeg"))


@app.route("/api/inbox/reports", methods=["GET"])
def inbox_list_reports():
    host = _get_request_host()
    can_own = _has_permission(host, "report.view_own")
    can_all = _has_permission(host, "report.view_all")
    if not can_own and not can_all and not _is_operator_session():
        return jsonify({"error": "Permission denied. Missing report view permission."}), 403
    rows = list_reports(str(RECEIPT_DB_PATH), host, limit=int(request.args.get("limit") or 100))
    q = (request.args.get("q") or "").strip().lower()
    mode = (request.args.get("mode") or "").strip().lower()
    creator_email = (request.args.get("creator_email") or "").strip().lower()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    if not _is_operator_session() and not can_all:
        me = (session.get("user_email") or "").strip().lower()
        rows = [x for x in rows if (x.get("creator_email") or "").strip().lower() == me]
    elif creator_email:
        rows = [x for x in rows if (x.get("creator_email") or "").strip().lower() == creator_email]

    if q:
        rows = [
            x for x in rows
            if q in (x.get("output_filename") or "").strip().lower()
            or q in (x.get("title") or "").strip().lower()
            or q in (x.get("employee_name") or "").strip().lower()
            or q in (x.get("department") or "").strip().lower()
        ]
    if mode:
        rows = [x for x in rows if (x.get("mode") or "").strip().lower() == mode]
    if date_from:
        rows = [x for x in rows if (x.get("created_at") or "")[:10] >= date_from]
    if date_to:
        rows = [x for x in rows if (x.get("created_at") or "")[:10] <= date_to]
    return jsonify({"reports": rows, "count": len(rows), "host": host})


@app.route("/api/inbox/reports/<int:report_id>", methods=["GET"])
def inbox_get_report_detail(report_id: int):
    host = _get_request_host()
    can_own = _has_permission(host, "report.view_own")
    can_all = _has_permission(host, "report.view_all")
    if not can_own and not can_all and not _is_operator_session():
        return jsonify({"error": "Permission denied. Missing report view permission."}), 403
    report = get_report(str(RECEIPT_DB_PATH), host, int(report_id))
    if not report:
        return jsonify({"error": "Report not found."}), 404
    if not _is_operator_session() and not can_all:
        me = (session.get("user_email") or "").strip().lower()
        if (report.get("creator_email") or "").strip().lower() != me:
            return jsonify({"error": "Cannot view another user's report."}), 403
    receipts = list_receipts_by_report_id(str(RECEIPT_DB_PATH), host, int(report_id), limit=3000)
    return jsonify({"host": host, "report": report, "receipts": receipts, "count": len(receipts)})


def _get_authorized_report_detail(host: str, report_id: int):
    can_own = _has_permission(host, "report.view_own")
    can_all = _has_permission(host, "report.view_all")
    if not can_own and not can_all and not _is_operator_session():
        return None, None, (jsonify({"error": "Permission denied. Missing report view permission."}), 403)
    report = get_report(str(RECEIPT_DB_PATH), host, int(report_id))
    if not report:
        return None, None, (jsonify({"error": "Report not found."}), 404)
    if not _is_operator_session() and not can_all:
        me = (session.get("user_email") or "").strip().lower()
        if (report.get("creator_email") or "").strip().lower() != me:
            return None, None, (jsonify({"error": "Cannot view another user's report."}), 403)
    receipts = list_receipts_by_report_id(str(RECEIPT_DB_PATH), host, int(report_id), limit=5000)
    return report, receipts, None


@app.route("/api/inbox/reports/<int:report_id>/line-items.csv", methods=["GET"])
def inbox_export_report_line_items_csv(report_id: int):
    host = _get_request_host()
    report, receipts, err = _get_authorized_report_detail(host, report_id)
    if err:
        return err
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["receipt_id", "file", "date", "merchant", "amount", "currency", "category", "memo", "status", "lifecycle_state"])
    for r in receipts:
        writer.writerow([
            r.get("id"),
            r.get("orig_filename") or "",
            r.get("date") or "",
            r.get("merchant") or "",
            int(r.get("amount") or 0),
            r.get("currency") or "",
            r.get("category") or "",
            r.get("memo") or "",
            r.get("status") or "",
            r.get("lifecycle_state") or "",
        ])
    filename = f"ReportLineItems_{int(report.get('id') or report_id)}.csv"
    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.route("/api/inbox/reports/<int:report_id>/line-items.xlsx", methods=["GET"])
def inbox_export_report_line_items_xlsx(report_id: int):
    host = _get_request_host()
    report, receipts, err = _get_authorized_report_detail(host, report_id)
    if err:
        return err
    wb = Workbook()
    ws = wb.active
    ws.title = "LineItems"
    ws.append(["receipt_id", "file", "date", "merchant", "amount", "currency", "category", "memo", "status", "lifecycle_state"])
    for r in receipts:
        ws.append([
            int(r.get("id") or 0),
            r.get("orig_filename") or "",
            r.get("date") or "",
            r.get("merchant") or "",
            int(r.get("amount") or 0),
            r.get("currency") or "",
            r.get("category") or "",
            r.get("memo") or "",
            r.get("status") or "",
            r.get("lifecycle_state") or "",
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"ReportLineItems_{int(report.get('id') or report_id)}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/inbox/reports/generate", methods=["POST"])
def inbox_generate_report():
    body = request.get_json(silent=True) or {}
    host = _get_request_host()
    can_generate = _has_permission(host, "report.generate")
    can_submit = _has_permission(host, "report.submit")
    if not (can_generate or can_submit):
        return jsonify({"error": "Permission denied. Missing report generate permission."}), 403
    receipt_ids = body.get("receipt_ids") or []
    rows = []

    if isinstance(receipt_ids, list) and receipt_ids:
        for rid in receipt_ids:
            try:
                item = get_receipt(str(RECEIPT_DB_PATH), host, int(rid))
            except Exception:
                item = None
            if item:
                rows.append(item)
    else:
        filters = body.get("filters") if isinstance(body.get("filters"), dict) else {}
        min_amount = filters.get("min_amount")
        max_amount = filters.get("max_amount")
        try:
            min_amount = int(min_amount) if min_amount is not None and str(min_amount) != "" else None
            max_amount = int(max_amount) if max_amount is not None and str(max_amount) != "" else None
        except Exception:
            return jsonify({"error": "min_amount/max_amount must be integers"}), 400
        rows = list_receipts_for_report(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            date_from=(filters.get("date_from") or None),
            date_to=(filters.get("date_to") or None),
            category=(filters.get("category") or None),
            merchant=(filters.get("merchant") or None),
            min_amount=min_amount,
            max_amount=max_amount,
            only_unassigned=bool(filters.get("only_unassigned", True)),
            status=(filters.get("status") or None),
            limit=2000,
        )
    if not rows:
        return jsonify({"error": "No valid receipts found for this host."}), 400

    can_view_all_reports = _is_operator_session() or _has_permission(host, "report.view_all")
    if not can_view_all_reports:
        uid = int(session.get("user_id") or 0)
        rows = [r for r in rows if int(r.get("uploader_user_id") or 0) == uid]
        if not rows:
            return jsonify({"error": "You can submit reports only for your own receipts."}), 403

    account_map = _get_effective_account_map(host)
    default_currency = str(body.get("currency") or "USD").strip().upper() or "USD"
    receipts = [
        {
            "filename": r.get("orig_filename"),
            "date": r.get("date"),
            "merchant": r.get("merchant"),
            "amount": int(r.get("amount") or 0),
            "currency": (r.get("currency") or default_currency),
            "category": r.get("category") or "MISCELLANEOUS",
            "memo": r.get("memo") or "",
            "account_code": account_map.get((r.get("category") or "").upper(), ""),
        }
        for r in rows
    ]
    if not any(parse_date(r.get("date")) and int(r.get("amount") or 0) > 0 for r in receipts):
        return jsonify({"error": "At least one receipt must have a valid date and amount."}), 400

    mode = (body.get("mode") or "domestic").strip().lower()
    template_path = get_template_path(mode, host=host)
    if not template_path:
        return jsonify({"error": f"'{mode}' template is missing."}), 400

    cfg = load_config(host)
    exchange_rate = float(body.get("exchange_rate") or DEFAULT_EXCHANGE_RATE)
    exchange_rates_raw = body.get("exchange_rates")
    exchange_rates = None
    if exchange_rates_raw and isinstance(exchange_rates_raw, dict):
        parsed_rates = {}
        for k, v in exchange_rates_raw.items():
            key = str(k or "").strip()
            if not key or v in (None, ""):
                continue
            if isinstance(v, dict):
                nested = {}
                for nk, nv in v.items():
                    nkey = str(nk or "").strip().upper()
                    if not nkey or nv in (None, ""):
                        continue
                    try:
                        nested[nkey] = float(nv)
                    except Exception:
                        continue
                if nested:
                    parsed_rates[key] = nested
            else:
                try:
                    parsed_rates[key] = float(v)
                except Exception:
                    continue
        exchange_rates = parsed_rates or None

    employee_info = {
        "name": body.get("employee_name", ""),
        "department": body.get("department", ""),
        "employee_id": body.get("employee_id", ""),
        "manager": body.get("manager", ""),
        "project": body.get("project", ""),
    }
    trip_title = str(body.get("trip_title") or body.get("title") or "").strip()
    submission_date = parse_date(body.get("submission_date"))
    settlement_month = body.get("settlement_month")
    period_mode = str(body.get("period_mode") or "manual").strip().lower()
    if period_mode not in {"manual", "auto"}:
        period_mode = "manual"
    period_from = str(body.get("period_from") or "").strip()
    period_to = str(body.get("period_to") or "").strip()
    if period_mode == "auto":
        valid_dates = sorted(
            {
                str(r.get("date") or "").strip()
                for r in rows
                if parse_date(r.get("date"))
            }
        )
        if valid_dates:
            period_from = valid_dates[0]
            period_to = valid_dates[-1]

    active_types = [
        {"id": str(t["id"]), "label": str(t.get("label") or t["id"])}
        for t in cfg.get("expense_types", [])
        if t.get("enabled") and t.get("id")
    ]

    session_id = str(uuid.uuid4())[:8]
    output_filename = f"ExpenseReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = OUTPUT_FOLDER / session_id
    output_path.mkdir(exist_ok=True)
    output_file = output_path / output_filename

    try:
        date_range = None
        if parse_date(period_from) and parse_date(period_to):
            date_range = (parse_date(period_from), parse_date(period_to))
        fill_expense_report(
            template_path=str(template_path),
            receipts=receipts,
            output_path=str(output_file),
            mode=mode,
            employee_info=employee_info,
            trip_title=trip_title,
            exchange_rate=exchange_rate,
            exchange_rates=exchange_rates,
            date_range=date_range,
            submission_date=submission_date,
            settlement_month=int(settlement_month) if settlement_month else None,
            active_types=active_types,
            company_info=cfg.get("company", {}),
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    total_amount = sum(int(r.get("amount") or 0) for r in rows)
    report_row = create_report(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        creator_user_id=int(session.get("user_id")) if session.get("user_id") else None,
        creator_email=(session.get("user_email") or ""),
        title=str(body.get("title") or "").strip(),
        mode=mode,
        receipt_count=len(rows),
        total_amount=float(total_amount),
        currency=default_currency,
        output_session_id=session_id,
        output_filename=output_filename,
        employee_name=str(body.get("employee_name") or ""),
        department=str(body.get("department") or ""),
        employee_id=str(body.get("employee_id") or ""),
        manager=str(body.get("manager") or ""),
        project=str(body.get("project") or ""),
        period_from=period_from,
        period_to=period_to,
        trip_purpose=str(body.get("trip_purpose") or ""),
        notes=str(body.get("notes") or ""),
    )
    for row in rows:
        update_receipt(
            db_path=str(RECEIPT_DB_PATH),
            host=host,
            receipt_id=int(row["id"]),
            fields={
                "report_status": "assigned",
                "report_id": report_row["id"],
                "status": "processed",
                "lifecycle_state": "ASSIGNED_TO_REPORT",
            },
        )
    return jsonify(
        {
            "status": "ok",
            "report": report_row,
            "download_url": f"/api/download/{session_id}/{output_filename}",
        }
    )


# ─── Receipt Upload ───

@app.route("/api/upload-receipts", methods=["POST"])
def upload_receipts():
    if "receipts" not in request.files:
        return jsonify({"error": "No receipt files provided."}), 400

    session_id = str(uuid.uuid4())[:8]
    session_dir = UPLOAD_FOLDER / session_id
    session_dir.mkdir(exist_ok=True)

    uploaded = []
    for file in request.files.getlist("receipts"):
        if file and allowed_file(file.filename, ALLOWED_IMAGE_EXT):
            filename = secure_filename(file.filename)
            file.save(str(session_dir / filename))
            uploaded.append(filename)

    if not uploaded:
        return jsonify({"error": "No valid image files found."}), 400

    def sort_key(name):
        stem = Path(name).stem
        return int(stem) if stem.isdigit() else stem

    uploaded.sort(key=sort_key)
    return jsonify({"session_id": session_id, "files": uploaded, "count": len(uploaded)})


# ─── OCR ───

@app.route("/api/ocr/<session_id>/<filename>", methods=["POST"])
def ocr_single(session_id, filename):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        return jsonify({"error": "서버에 ANTHROPIC_API_KEY가 설정되지 않았습니다. 관리자에게 문의하세요."}), 500

    image_path = UPLOAD_FOLDER / session_id / secure_filename(filename)
    if not image_path.exists():
        return jsonify({"error": "File not found."}), 404

    try:
        result = process_receipt_image(str(image_path), api_key)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e), "filename": filename}), 500


# ─── Serve receipt image ───

@app.route("/api/image/<session_id>/<filename>")
def serve_receipt_image(session_id, filename):
    image_path = UPLOAD_FOLDER / secure_filename(session_id) / secure_filename(filename)
    if not image_path.exists():
        return jsonify({"error": "Image not found."}), 404
    ext = image_path.suffix.lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                ".webp": "image/webp", ".heic": "image/heic"}
    return send_file(str(image_path), mimetype=mime_map.get(ext, "image/jpeg"))


# ─── Generate Report ───

@app.route("/api/generate/<session_id>", methods=["POST"])
def generate_report(session_id):
    data = request.json
    receipts = data.get("receipts", [])
    if not receipts:
        return jsonify({"error": "No receipt data provided."}), 400

    # 모드: domestic / international
    mode = data.get("mode", "domestic")
    template_path = get_template_path(mode, host=host)
    if not template_path:
        return jsonify({
            "error": f"'{mode}' 템플릿이 없습니다. 관리자 페이지(/admin)에서 템플릿을 업로드하세요."
        }), 400

    cfg = load_config(_get_request_host())

    # 환율
    exchange_rate      = float(data.get("exchange_rate") or DEFAULT_EXCHANGE_RATE)
    exchange_rates_raw = data.get("exchange_rates")
    exchange_rates     = None
    if exchange_rates_raw and isinstance(exchange_rates_raw, dict):
        parsed_rates = {}
        for k, v in exchange_rates_raw.items():
            key = str(k or "").strip()
            if not key or v in (None, ""):
                continue
            if isinstance(v, dict):
                nested = {}
                for nk, nv in v.items():
                    nkey = str(nk or "").strip().upper()
                    if not nkey or nv in (None, ""):
                        continue
                    try:
                        nested[nkey] = float(nv)
                    except Exception:
                        continue
                if nested:
                    parsed_rates[key] = nested
            else:
                try:
                    parsed_rates[key] = float(v)
                except Exception:
                    continue
        exchange_rates = parsed_rates or None

    # 기간 필터
    date_range = None
    dr = data.get("date_range")
    if dr and len(dr) == 2 and dr[0] and dr[1]:
        start_d = parse_date(dr[0])
        end_d   = parse_date(dr[1])
        if start_d and end_d:
            date_range = (start_d, end_d)

    # 직원 정보
    employee_info = {
        "name":       data.get("employee_name", ""),
        "department": data.get("department", ""),
        "employee_id":data.get("employee_id", ""),
        "manager":    data.get("manager", ""),
        "project":    data.get("project", ""),
    }
    trip_title        = data.get("trip_title", "")
    submission_date   = parse_date(data.get("submission_date"))
    settlement_month  = data.get("settlement_month")

    # 활성화된 expense type 목록
    active_types = [
        {"id": str(t["id"]), "label": str(t.get("label") or t["id"])}
        for t in cfg.get("expense_types", [])
        if t.get("enabled") and t.get("id")
    ]

    output_filename = f"ExpenseReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path     = OUTPUT_FOLDER / session_id
    output_path.mkdir(exist_ok=True)
    output_file     = output_path / output_filename

    try:
        fill_expense_report(
            template_path   = str(template_path),
            receipts        = receipts,
            output_path     = str(output_file),
            mode            = mode,
            employee_info   = employee_info,
            trip_title      = trip_title,
            exchange_rate   = exchange_rate,
            exchange_rates  = exchange_rates,
            date_range      = date_range,
            submission_date = submission_date,
            settlement_month= int(settlement_month) if settlement_month else None,
            active_types    = active_types,
            company_info    = cfg.get("company", {}),
        )
        total_amount = sum(int(r.get("amount") or 0) for r in receipts if isinstance(r, dict))
        create_report(
            db_path=str(RECEIPT_DB_PATH),
            host=_get_request_host(),
            creator_user_id=int(session.get("user_id")) if session.get("user_id") else None,
            creator_email=(session.get("user_email") or ""),
            title=str(data.get("title") or "Legacy Workflow").strip(),
            mode=mode,
            receipt_count=len(receipts),
            total_amount=float(total_amount),
            currency=str(data.get("currency") or ""),
            output_session_id=session_id,
            output_filename=output_filename,
        )
        return jsonify({
            "status":       "ok",
            "filename":     output_filename,
            "download_url": f"/api/download/{session_id}/{output_filename}",
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/download/<session_id>/<filename>")
def download_file(session_id, filename):
    host = _get_request_host()
    safe_sid = secure_filename(session_id)
    safe_name = secure_filename(filename)
    report = get_report_by_output(str(RECEIPT_DB_PATH), host, safe_sid, safe_name)
    if not report and not _is_operator_session():
        return jsonify({"error": "Download metadata not found for this host."}), 403
    if report and not _is_operator_session():
        can_all = _has_permission(host, "report.view_all")
        me = (session.get("user_email") or "").strip().lower()
        owner = (report.get("creator_email") or "").strip().lower()
        if not can_all and me != owner:
            return jsonify({"error": "You are not allowed to download this report."}), 403

    file_path = OUTPUT_FOLDER / safe_sid / safe_name
    if not file_path.exists():
        return jsonify({"error": "File not found."}), 404
    return send_file(str(file_path), as_attachment=True, download_name=safe_name)


# ─────────────────────────────────────────────
# SSO Routes
# ─────────────────────────────────────────────

@app.route("/auth/login/microsoft")
def auth_login_microsoft():
    if not _is_sso_configured():
        return jsonify({"error": "Microsoft SSO is not fully configured."}), 400
    ms_client = oauth.create_client("microsoft")

    session["post_login_redirect"] = request.args.get("next", "/")
    redirect_uri = _effective_ms_redirect_uri()
    auth_params = {}
    if SSO_PROMPT:
        auth_params["prompt"] = SSO_PROMPT
    return ms_client.authorize_redirect(redirect_uri, **auth_params)


@app.route("/auth/login/local", methods=["POST"])
def auth_login_local():
    host = _get_request_host()
    local_enabled, local_password = _host_local_login_policy(host)
    if not local_enabled:
        return jsonify({"error": "Local password login is disabled."}), 400
    body = request.get_json(silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    name = (body.get("name") or "").strip()
    password = (body.get("password") or "").strip()
    if not email:
        return jsonify({"error": "Email is required."}), 400
    if not password or password != local_password:
        return jsonify({"error": "Invalid email/password."}), 403
    if not _is_email_allowed_for_host(email, host):
        return jsonify({"error": f"Email domain is not allowed for host '{host}'."}), 403
    is_admin = (
        email in EFFECTIVE_ADMIN_EMAILS
        or _is_operator_identity(email)
        or (not EFFECTIVE_ADMIN_EMAILS and password == ADMIN_PASSWORD)
    )

    user = upsert_user_by_email(
        db_path=str(USER_DB_PATH),
        email=email,
        name=name,
        is_admin=is_admin,
    )
    _set_session_user(user, provider="local")
    return jsonify({
        "status": "ok",
        "user": user,
        "redirect_url": _resolve_post_login_redirect(email, request.args.get("next", "/")),
    })


@app.route("/auth/callback/microsoft")
def auth_callback_microsoft():
    if not _is_sso_configured():
        return jsonify({"error": "Microsoft SSO is not configured."}), 500

    redirect_uri = _effective_ms_redirect_uri()
    code = request.args.get("code", "")
    if not code:
        return jsonify({"error": "Authorization code is missing."}), 400
    try:
        token = _exchange_ms_code_directly(code, redirect_uri)
    except Exception as e:
        return jsonify({"error": f"Microsoft token exchange failed: {str(e)}"}), 400

    access_token = token.get("access_token", "")
    if not access_token:
        return jsonify({"error": "No access token in Microsoft response."}), 400
    try:
        user_info = _fetch_ms_me(access_token)
    except Exception as e:
        return jsonify({"error": f"Could not read Microsoft profile (/me): {str(e)}"}), 400

    email = (
        user_info.get("email")
        or user_info.get("preferred_username")
        or user_info.get("userPrincipalName")
        or user_info.get("mail")
        or user_info.get("upn")
        or ""
    ).strip().lower()
    if not email:
        return jsonify({"error": "No email claim found in Microsoft token."}), 400
    host = _get_request_host()
    if not _is_email_allowed_for_host(email, host):
        return jsonify({"error": f"Email domain is not allowed for host '{host}'."}), 403

    provider_user_id = str(
        user_info.get("oid")
        or user_info.get("sub")
        or user_info.get("id")
        or ""
    ).strip()
    if not provider_user_id:
        return jsonify({"error": "No stable user id claim(oid/sub) found in token."}), 400

    org_name = _fetch_ms_org_name(access_token)
    user = upsert_user_from_oidc(
        db_path=str(USER_DB_PATH),
        provider="microsoft",
        provider_user_id=provider_user_id,
        email=email,
        name=str(user_info.get("name") or ""),
        tenant_id=str(user_info.get("tid") or ""),
        org_name=org_name,
        admin_emails=EFFECTIVE_ADMIN_EMAILS,
    )
    if _is_operator_identity(email) and not user.get("is_admin"):
        user = update_user_flags(str(USER_DB_PATH), user["id"], is_admin=True)
    _set_session_user(user, provider="microsoft")
    return_url = _resolve_post_login_redirect(email, session.pop("post_login_redirect", "/"))
    return redirect(return_url)


@app.route("/auth/me")
def auth_me():
    uid = session.get("user_id")
    if not uid:
        return jsonify({"authenticated": False})
    user = get_user_by_id(str(USER_DB_PATH), int(uid))
    if not user:
        _clear_session_user()
        return jsonify({"authenticated": False})
    host = _get_request_host()
    view = dict(user)
    view["is_admin"] = _is_admin_for_host(host)
    return jsonify({
        "authenticated": True,
        "provider": session.get("auth_provider"),
        "host": host,
        "operator": _is_operator_session(),
        "user": view,
    })


@app.route("/api/profile", methods=["GET", "PUT"])
def api_profile():
    uid = int(session.get("user_id") or 0)
    if uid <= 0:
        return jsonify({"error": "Login required."}), 401
    if request.method == "GET":
        user = get_user_by_id(str(USER_DB_PATH), uid)
        if not user:
            return jsonify({"error": "User not found."}), 404
        return jsonify(
            {
                "profile": {
                    "name": str(user.get("name") or ""),
                    "department": str(user.get("department") or ""),
                    "employee_id": str(user.get("employee_code") or ""),
                    "manager": str(user.get("manager_name") or ""),
                    "email": str(user.get("email") or ""),
                }
            }
        )
    body = request.get_json(silent=True) or {}
    user = update_user_profile(
        db_path=str(USER_DB_PATH),
        user_id=uid,
        name=body.get("name"),
        department=body.get("department"),
        employee_code=body.get("employee_id"),
        manager_name=body.get("manager"),
    )
    return jsonify(
        {
            "status": "ok",
            "profile": {
                "name": str(user.get("name") or ""),
                "department": str(user.get("department") or ""),
                "employee_id": str(user.get("employee_code") or ""),
                "manager": str(user.get("manager_name") or ""),
                "email": str(user.get("email") or ""),
            },
        }
    )


@app.route("/auth/logout", methods=["POST"])
def auth_logout():
    _clear_session_user()
    return jsonify({"status": "ok"})


@app.route("/logout")
def auth_logout_page():
    _clear_session_user()
    return redirect("/login")


# ─────────────────────────────────────────────
# Admin Routes
# ─────────────────────────────────────────────

def _check_admin(req) -> bool:
    """세션 인증 우선, 없으면 비밀번호 확인 후 세션 생성."""
    if _is_admin_authenticated(_get_request_host()):
        return True
    pw = _extract_admin_password(req)
    if pw and pw == ADMIN_PASSWORD:
        # Legacy password gate. Requires an existing authenticated user session.
        if _is_user_authenticated():
            session.permanent = True
            session["is_admin"] = True
            return True
        return False
    return False


@app.route("/admin")
def admin_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/admin")
    if not _is_admin_authenticated(_get_request_host()):
        return redirect("/")
    return render_template("admin.html")


@app.route("/platform")
def platform_page():
    if not _is_user_authenticated():
        return redirect("/login?next=/platform")
    if not _is_operator_session():
        return redirect("/admin")
    return render_template("platform.html")


@app.route("/admin/login", methods=["POST"])
def admin_login():
    if not _is_user_authenticated():
        return jsonify({"error": "Login required."}), 401
    pw = _extract_admin_password(request)
    if not pw:
        return jsonify({"error": "Password is required."}), 400
    if pw != ADMIN_PASSWORD:
        return jsonify({"error": "Invalid admin password."}), 403
    session.permanent = True
    session["is_admin"] = True
    return jsonify({"status": "ok"})


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    _clear_session_user()
    return jsonify({"status": "ok"})


@app.route("/admin/auth-status")
def admin_auth_status():
    missing = []
    if SSO_ENABLED and SSO_PROVIDER == "microsoft":
        if not MS_CLIENT_ID:
            missing.append("MS_CLIENT_ID")
        if not MS_CLIENT_SECRET:
            missing.append("MS_CLIENT_SECRET")
        elif _looks_like_secret_id(MS_CLIENT_SECRET):
            missing.append("MS_CLIENT_SECRET(Value)")
    return jsonify({
        "authenticated": _is_admin_authenticated(_get_request_host()),
        "provider": session.get("auth_provider"),
        "email": session.get("user_email"),
        "host": _get_request_host(),
        "operator": _is_operator_session(),
        "sso_enabled": SSO_ENABLED and SSO_PROVIDER == "microsoft",
        "sso_configured": _is_sso_configured(),
        "sso_missing": missing,
        "sso_callback": _effective_ms_redirect_uri(),
    })


@app.route("/admin/tenants", methods=["GET"])
def admin_list_tenants():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    rows = _load_tenants_registry()
    current_host = _get_request_host()
    known = {x.get("host") for x in rows}
    if current_host and current_host not in known:
        rows.append({"host": current_host, "name": current_host, "redirect_url": "", "active": True})
    if not _is_operator_session():
        rows = [x for x in rows if x.get("host") == current_host]
    return jsonify({"tenants": rows, "operator": _is_operator_session(), "current_host": current_host})


@app.route("/admin/tenants", methods=["POST"])
def admin_upsert_tenant():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    if not _is_operator_session():
        return jsonify({"error": "Only operator accounts can manage tenant registry."}), 403
    body = request.get_json(silent=True) or {}
    host = _normalize_host_value(str(body.get("host") or ""))
    if not host:
        return jsonify({"error": "host is required"}), 400
    name = str(body.get("name") or host).strip() or host
    redirect_url = str(body.get("redirect_url") or "").strip()
    active = bool(body.get("active", True))

    rows = _load_tenants_registry()
    updated = False
    for row in rows:
        if row.get("host") == host:
            row["name"] = name
            row["redirect_url"] = redirect_url
            row["active"] = active
            updated = True
            break
    if not updated:
        rows.append({"host": host, "name": name, "redirect_url": redirect_url, "active": active})
    _save_tenants_registry(rows)
    _ensure_tenant_known(host)
    return jsonify({"status": "ok", "tenant": {"host": host, "name": name, "redirect_url": redirect_url, "active": active}})


# ── Config GET / SAVE ──

@app.route("/admin/config", methods=["GET"])
def admin_get_config():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    return jsonify({**load_config(host), "_config_host": host})


@app.route("/admin/config", methods=["POST"])
def admin_save_config():
    if not _check_admin(request):
        return jsonify({"error": "Invalid admin password."}), 403

    body = request.json or {}
    host = _resolve_target_host(request, body)
    _ensure_tenant_known(host)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    if "company" in body:
        ok, err = _require_any_permission(host, ["company.manage"])
        if not ok:
            return jsonify({"error": err}), 403
    if "expense_types" in body:
        ok, err = _require_any_permission(host, ["category.manage"])
        if not ok:
            return jsonify({"error": err}), 403
    if "auth" in body:
        ok, err = _require_any_permission(host, ["policy.manage"])
        if not ok:
            return jsonify({"error": err}), 403
    if "wizard" in body:
        ok, err = _require_any_permission(host, ["policy.manage"])
        if not ok:
            return jsonify({"error": err}), 403
    cfg  = load_config(host)

    # company
    if "company" in body:
        cfg["company"].update({
            k: v for k, v in body["company"].items()
            if k in cfg["company"] and k != "logo_filename"
        })

    # fields
    if "fields" in body:
        for fid, fdata in body["fields"].items():
            if fid in cfg["fields"]:
                cfg["fields"][fid].update({
                    k: v for k, v in fdata.items() if k in ("enabled", "label")
                })

    # expense_types — 전체 교체 (순서 포함)
    if "expense_types" in body:
        types = []
        for t in body["expense_types"]:
            if isinstance(t, dict) and t.get("id") and t.get("label"):
                types.append({
                    "id":      str(t["id"]).upper().replace(" ", "_")[:30],
                    "label":   str(t["label"])[:60],
                    "enabled": bool(t.get("enabled", True)),
                })
        if types:
            cfg["expense_types"] = types

    # modes
    if "modes" in body:
        for mid, mdata in body["modes"].items():
            if mid in cfg["modes"]:
                cfg["modes"][mid]["enabled"] = bool(mdata.get("enabled", True))

    # login page
    if "login_page" in body and isinstance(body["login_page"], dict):
        lp = body["login_page"]
        if "interval_sec" in lp:
            try:
                sec = int(lp["interval_sec"])
                cfg["login_page"]["interval_sec"] = min(max(sec, 3), 120)
            except Exception:
                pass

    # auth
    if "auth" in body and isinstance(body["auth"], dict):
        ab = body["auth"]
        if "allowed_email_domains" in ab:
            cfg["auth"]["allowed_email_domains"] = sorted(_normalize_domain_set(ab.get("allowed_email_domains")))
        if "admin_emails" in ab:
            cfg["auth"]["admin_emails"] = sorted(_normalize_email_set(ab.get("admin_emails")))
        if "local_login_enabled" in ab:
            cfg["auth"]["local_login_enabled"] = bool(ab.get("local_login_enabled"))
        if "local_login_password" in ab:
            pw = str(ab.get("local_login_password") or "").strip()
            if pw:
                cfg["auth"]["local_login_password"] = pw

    # wizard
    if "wizard" in body and isinstance(body["wizard"], dict):
        wb = body["wizard"]
        if "risk_priority" in wb and isinstance(wb["risk_priority"], list):
            allowed = {"OCR_FAILED", "IMAGE_UNREADABLE", "LOW_CONFIDENCE", "NEEDS_REVIEW"}
            rp = [str(x).strip().upper() for x in wb["risk_priority"] if str(x).strip().upper() in allowed]
            if rp:
                cfg.setdefault("wizard", {})
                cfg["wizard"]["risk_priority"] = rp
        if "cache_ttl_hours" in wb:
            try:
                ttl = int(wb.get("cache_ttl_hours"))
                cfg.setdefault("wizard", {})
                cfg["wizard"]["cache_ttl_hours"] = max(1, min(ttl, 24 * 30))
            except Exception:
                pass

    save_config(cfg, host)
    return jsonify({"status": "ok", "config": cfg})


# ── User Management ──

@app.route("/admin/users", methods=["GET"])
def admin_list_users():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    allowed_domains, tenant_admins = _get_host_auth_settings(host)
    scope_platform = (request.args.get("scope", "") or "").strip().lower() == "platform"
    # /admin is tenant-scoped by default. Cross-tenant visibility is only for explicit /platform scope.
    operator_session = _is_operator_session() if scope_platform else False
    users = list_users(str(USER_DB_PATH))
    visible = []
    for u in users:
        email = (u.get("email") or "").strip().lower()
        if not _is_user_visible_for_host(email, host, allowed_domains, tenant_admins, operator_session):
            continue
        item = dict(u)
        item["is_admin"] = (_is_operator_identity(email) or (email in tenant_admins)) if tenant_admins else bool(u.get("is_admin"))
        visible.append(item)
    return jsonify({"users": visible, "_target_host": host, "_tenant_scoped": bool(tenant_admins)})


@app.route("/admin/users", methods=["POST"])
def admin_create_or_update_user():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    body = request.get_json(silent=True) or {}
    email = (body.get("email") or "").strip().lower()
    name = (body.get("name") or "").strip()
    is_admin = bool(body.get("is_admin", True))
    host = _resolve_target_host(request, body)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    if not email:
        return jsonify({"error": "email is required"}), 400
    try:
        # Tenant admin is managed in host config. Keep operator/global admins in DB for backward compatibility.
        db_admin = bool(is_admin and (_is_operator_identity(email) or email in EFFECTIVE_ADMIN_EMAILS))
        user = upsert_user_by_email(
            db_path=str(USER_DB_PATH),
            email=email,
            name=name,
            is_admin=db_admin,
        )
        _set_tenant_admin_flag(host, email, is_admin)
        user_view = dict(user)
        user_view["is_admin"] = is_admin
        return jsonify({"status": "ok", "user": user_view})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/admin/users/<int:user_id>", methods=["PATCH"])
def admin_update_user(user_id: int):
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    body = request.get_json(silent=True) or {}
    is_admin = body["is_admin"] if "is_admin" in body else None
    is_active = body["is_active"] if "is_active" in body else None
    host = _resolve_target_host(request, body)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    target = get_user_by_id(str(USER_DB_PATH), user_id)
    if not target:
        return jsonify({"error": "user not found"}), 404

    # 현재 로그인한 사용자가 본인 admin 권한을 내리는 것은 방지
    session_user_id = session.get("user_id")
    if session_user_id and int(session_user_id) == int(user_id) and is_admin is False:
        return jsonify({"error": "You cannot revoke your own admin role."}), 400

    # 운영자 계정은 admin/active 강등 불가(옵션)
    target_email = (target.get("email") or "").lower()
    if PROTECT_OPERATOR_ACCOUNTS and _is_operator_identity(target_email) and (is_admin is False or is_active is False):
        return jsonify({"error": "Operator account cannot be demoted or deactivated."}), 400

    try:
        if is_admin is not None:
            _set_tenant_admin_flag(host, target_email, bool(is_admin))
        # DB is_admin flag is reserved for global/operator compatibility.
        user = target
        if is_active is not None:
            user = update_user_flags(
                db_path=str(USER_DB_PATH),
                user_id=user_id,
                is_active=is_active,
            )
        _, tenant_admins = _get_host_auth_settings(host)
        user_view = dict(user)
        user_view["is_admin"] = _is_operator_identity(target_email) or (target_email in tenant_admins) if tenant_admins else bool(user.get("is_admin"))
        return jsonify({"status": "ok", "user": user_view})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/admin/user-permissions", methods=["GET"])
def admin_list_user_permissions():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["policy.manage", "user.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    allowed_domains, tenant_admins = _get_host_auth_settings(host)
    scope_platform = (request.args.get("scope", "") or "").strip().lower() == "platform"
    # /admin is tenant-scoped by default. Cross-tenant visibility is only for explicit /platform scope.
    operator_session = _is_operator_session() if scope_platform else False
    users = list_users(str(USER_DB_PATH))
    explicit_rows = list_user_host_permissions(str(USER_DB_PATH), host)
    explicit_map = {int(r["user_id"]): set(r.get("permissions") or []) for r in explicit_rows}

    items = []
    for u in users:
        email = (u.get("email") or "").strip().lower()
        if not _is_user_visible_for_host(email, host, allowed_domains, tenant_admins, operator_session):
            continue
        explicit = sorted(explicit_map.get(int(u["id"]), set()))
        if _is_operator_identity(email) or email in tenant_admins:
            effective = sorted(DEFAULT_ADMIN_PERMISSIONS)
        else:
            effective = sorted(explicit_map.get(int(u["id"]), set(DEFAULT_GENERAL_PERMISSIONS)))
        items.append(
            {
                "id": int(u["id"]),
                "email": email,
                "name": u.get("name") or "",
                "explicit_permissions": explicit,
                "effective_permissions": effective,
                "is_tenant_admin": email in tenant_admins,
                "is_operator": _is_operator_identity(email),
            }
        )
    return jsonify({"host": host, "available_permissions": ALL_PERMISSION_KEYS, "users": items})


@app.route("/admin/user-permissions/<int:user_id>", methods=["PATCH"])
def admin_update_user_permissions(user_id: int):
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    body = request.get_json(silent=True) or {}
    host = _resolve_target_host(request, body)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["policy.manage", "user.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    target = get_user_by_id(str(USER_DB_PATH), user_id)
    if not target:
        return jsonify({"error": "user not found"}), 404
    perms = body.get("permissions")
    if not isinstance(perms, list):
        return jsonify({"error": "permissions(list) is required"}), 400

    normalized = sorted({str(x).strip() for x in perms if str(x).strip() in ALL_PERMISSION_KEYS})
    assigned = set_user_host_permissions(str(USER_DB_PATH), int(user_id), host, normalized)
    return jsonify({"status": "ok", "host": host, "user_id": int(user_id), "permissions": assigned})


@app.route("/admin/inbox/receipts", methods=["GET"])
def admin_inbox_receipts():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    status = (request.args.get("status") or "").strip() or None
    lifecycle_state = (request.args.get("lifecycle_state") or "").strip().upper() or None
    report_status = (request.args.get("report_status") or "").strip() or None
    q = (request.args.get("q") or "").strip() or None
    period_days = int(request.args.get("period_days") or 0)
    user_email = (request.args.get("user_email") or "").strip().lower()
    items = list_receipts(
        db_path=str(RECEIPT_DB_PATH),
        host=host,
        status=status,
        lifecycle_state=lifecycle_state,
        report_status=report_status,
        search=q,
        limit=int(request.args.get("limit") or 500),
    )
    if user_email:
        items = [x for x in items if (x.get("uploader_email") or "").strip().lower() == user_email]
    if period_days > 0:
        items = [x for x in items if _within_last_days(x.get("created_at") or x.get("updated_at"), period_days)]
    return jsonify({"host": host, "receipts": items, "count": len(items)})


@app.route("/admin/inbox/reports", methods=["GET"])
def admin_inbox_reports():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    period_days = int(request.args.get("period_days") or 0)
    user_email = (request.args.get("user_email") or "").strip().lower()
    rows = list_reports(str(RECEIPT_DB_PATH), host, limit=int(request.args.get("limit") or 300))
    if user_email:
        rows = [x for x in rows if (x.get("creator_email") or "").strip().lower() == user_email]
    if period_days > 0:
        rows = [x for x in rows if _within_last_days(x.get("created_at"), period_days)]
    return jsonify({"host": host, "reports": rows, "count": len(rows)})


@app.route("/admin/inbox/kpi-by-user", methods=["GET"])
def admin_inbox_kpi_by_user():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    period_days = int(request.args.get("period_days") or 0)
    receipts = list_receipts(str(RECEIPT_DB_PATH), host, limit=int(request.args.get("limit") or 5000))
    reports = list_reports(str(RECEIPT_DB_PATH), host, limit=int(request.args.get("limit_reports") or 3000))
    if period_days > 0:
        receipts = [x for x in receipts if _within_last_days(x.get("created_at") or x.get("updated_at"), period_days)]
        reports = [x for x in reports if _within_last_days(x.get("created_at"), period_days)]
    by_user = {}
    for r in receipts:
        email = (r.get("uploader_email") or "").strip().lower() or "(unknown)"
        slot = by_user.setdefault(email, {"email": email, "receipts": 0, "needs_review": 0, "duplicate": 0, "assigned": 0, "reports": 0})
        slot["receipts"] += 1
        if (r.get("status") or "") == "needs_review":
            slot["needs_review"] += 1
        if (r.get("status") or "") == "duplicate":
            slot["duplicate"] += 1
        if (r.get("report_status") or "") == "assigned":
            slot["assigned"] += 1
    for rp in reports:
        email = (rp.get("creator_email") or "").strip().lower() or "(unknown)"
        slot = by_user.setdefault(email, {"email": email, "receipts": 0, "needs_review": 0, "duplicate": 0, "assigned": 0, "reports": 0})
        slot["reports"] += 1
    rows = sorted(by_user.values(), key=lambda x: (-int(x["receipts"]), x["email"]))
    return jsonify({"host": host, "users": rows, "count": len(rows)})


@app.route("/admin/inbox/audit-logs", methods=["GET"])
def admin_inbox_audit_logs():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    user_email = (request.args.get("user_email") or "").strip().lower() or None
    period_days = int(request.args.get("period_days") or 0)
    receipt_id_raw = (request.args.get("receipt_id") or "").strip()
    receipt_id = int(receipt_id_raw) if receipt_id_raw.isdigit() else None
    rows = list_receipt_audit_logs_by_host(
        str(RECEIPT_DB_PATH),
        host,
        user_email=user_email,
        receipt_id=receipt_id,
        limit=int(request.args.get("limit") or 500),
    )
    if period_days > 0:
        rows = [x for x in rows if _within_last_days(x.get("created_at"), period_days)]
    return jsonify({"host": host, "logs": rows, "count": len(rows)})


@app.route("/admin/inbox/audit-logs.csv", methods=["GET"])
def admin_inbox_audit_logs_csv():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    user_email = (request.args.get("user_email") or "").strip().lower() or None
    period_days = int(request.args.get("period_days") or 0)
    receipt_id_raw = (request.args.get("receipt_id") or "").strip()
    receipt_id = int(receipt_id_raw) if receipt_id_raw.isdigit() else None
    rows = list_receipt_audit_logs_by_host(
        str(RECEIPT_DB_PATH),
        host,
        user_email=user_email,
        receipt_id=receipt_id,
        limit=int(request.args.get("limit") or 2000),
    )
    if period_days > 0:
        rows = [x for x in rows if _within_last_days(x.get("created_at"), period_days)]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["created_at", "receipt_id", "actor_email", "action", "changed_fields", "diff_text", "before_json", "after_json"])
    for x in rows:
        diff_text = ""
        try:
            fields = json.loads(x.get("changed_fields") or "[]")
            before = json.loads(x.get("before_json") or "{}")
            after = json.loads(x.get("after_json") or "{}")
            if isinstance(fields, list):
                chunks = []
                for f in fields:
                    key = str(f)
                    chunks.append(f"{key}: {before.get(key, '')} -> {after.get(key, '')}")
                diff_text = " | ".join(chunks)
        except Exception:
            diff_text = ""
        writer.writerow([
            x.get("created_at") or "",
            x.get("receipt_id") or "",
            x.get("actor_email") or "",
            x.get("action") or "",
            x.get("changed_fields") or "[]",
            diff_text,
            x.get("before_json") or "{}",
            x.get("after_json") or "{}",
        ])
    filename = f"audit_logs_{host}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.route("/admin/inbox/shared-preset-audit-logs", methods=["GET"])
def admin_shared_preset_audit_logs():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage", "policy.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    preset_name = (request.args.get("preset_name") or "").strip() or None
    period_days = int(request.args.get("period_days") or 0)
    rows = list_shared_inbox_filter_preset_audit_logs(
        str(USER_DB_PATH),
        host,
        preset_name=preset_name,
        limit=int(request.args.get("limit") or 300),
    )
    if period_days > 0:
        rows = [x for x in rows if _within_last_days(x.get("created_at"), period_days)]
    return jsonify({"host": host, "logs": rows, "count": len(rows)})


@app.route("/admin/inbox/shared-preset-audit-logs.csv", methods=["GET"])
def admin_shared_preset_audit_logs_csv():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _resolve_target_host(request)
    ok, err = _ensure_tenant_admin_allowed(host)
    if not ok:
        return jsonify({"error": err}), 403
    ok, err = _require_any_permission(host, ["user.manage", "policy.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    preset_name = (request.args.get("preset_name") or "").strip() or None
    period_days = int(request.args.get("period_days") or 0)
    rows = list_shared_inbox_filter_preset_audit_logs(
        str(USER_DB_PATH),
        host,
        preset_name=preset_name,
        limit=int(request.args.get("limit") or 1000),
    )
    if period_days > 0:
        rows = [x for x in rows if _within_last_days(x.get("created_at"), period_days)]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["created_at", "preset_name", "actor_email", "action", "before_json", "after_json"])
    for x in rows:
        writer.writerow([
            x.get("created_at") or "",
            x.get("preset_name") or "",
            x.get("actor_email") or "",
            x.get("action") or "",
            x.get("before_json") or "{}",
            x.get("after_json") or "{}",
        ])
    filename = f"shared_preset_audit_logs_{host}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(out.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── Login Page Slides ──

@app.route("/admin/login-slides", methods=["GET"])
def admin_login_slides():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    cfg = load_config(_get_request_host())
    return jsonify({
        "interval_sec": int(cfg.get("login_page", {}).get("interval_sec") or 8),
        "slides": build_login_slides(cfg),
    })


@app.route("/admin/login-slides", methods=["POST"])
def admin_upload_login_slide():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    if "image" not in request.files:
        return jsonify({"error": "No image file provided."}), 400
    file = request.files["image"]
    if not allowed_file(file.filename, ALLOWED_LOGO_EXT):
        return jsonify({"error": "Accepted formats: PNG, JPG, GIF, SVG, WEBP"}), 400

    ext = file.filename.rsplit(".", 1)[1].lower()
    filename = f"{uuid.uuid4().hex[:12]}.{ext}"
    save_path = LOGIN_BG_FOLDER / filename
    file.save(str(save_path))

    host = _get_request_host()
    cfg = load_config(host)
    cfg.setdefault("login_page", {}).setdefault("slides", [])
    cfg["login_page"]["slides"].append({
        "filename": filename,
        "caption": (request.form.get("caption", "") or "")[:120],
    })
    save_config(cfg, host)
    return jsonify({"status": "ok", "slide": {"filename": filename, "url": f"/static/login_bg/{filename}"}})


@app.route("/admin/login-slides/<filename>", methods=["DELETE"])
def admin_delete_login_slide(filename):
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    safe_name = secure_filename(filename)
    host = _get_request_host()
    cfg = load_config(host)
    slides = cfg.get("login_page", {}).get("slides", [])
    cfg["login_page"]["slides"] = [
        s for s in slides
        if str(s.get("filename", "")) != safe_name
    ]
    p = LOGIN_BG_FOLDER / safe_name
    if p.exists():
        try:
            p.unlink()
        except Exception:
            pass
    save_config(cfg, host)
    return jsonify({"status": "ok"})


# ── Logo Upload ──

@app.route("/admin/upload-logo", methods=["POST"])
def admin_upload_logo():
    if not _check_admin(request):
        return jsonify({"error": "Invalid admin password."}), 403
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["company.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    if "logo" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    file = request.files["logo"]
    if not allowed_file(file.filename, ALLOWED_LOGO_EXT):
        return jsonify({"error": "Accepted formats: PNG, JPG, GIF, SVG, WEBP"}), 400

    host_key  = _host_to_config_key(host)
    ext       = file.filename.rsplit(".", 1)[1].lower()
    logo_name = f"company_logo_{host_key}.{ext}"
    save_path = LOGO_FOLDER / logo_name
    file.save(str(save_path))

    cfg = load_config(host)
    cfg["company"]["logo_filename"] = logo_name
    save_config(cfg, host)

    return jsonify({
        "status":   "ok",
        "filename": logo_name,
        "url":      f"/static/logos/{logo_name}",
    })


# ── Template Upload ──

@app.route("/admin/upload-template", methods=["POST"])
def admin_upload_template():
    if not _check_admin(request):
        return jsonify({"error": "Invalid admin password."}), 403
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["template.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    if "template" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    file = request.files["template"]
    if not allowed_file(file.filename, ALLOWED_EXCEL_EXT):
        return jsonify({"error": "Only .xlsx files are accepted."}), 400

    mode = request.form.get("mode", "domestic")
    fname = _template_filename_for_mode(mode)
    save_path = _tenant_template_dir(host) / fname
    file.save(str(save_path))
    stat = save_path.stat()
    return jsonify({
        "status":   "ok",
        "filename": fname,
        "size":     stat.st_size,
        "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
    })


@app.route("/admin/template-info")
def admin_template_info():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _get_request_host()
    result = {}
    for mode, fname in [("domestic", "template_domestic.xlsx"),
                         ("international", "template_international.xlsx"),
                         ("default", "template.xlsx")]:
        tenant_p = _tenant_template_dir(host) / fname
        default_p = DEFAULT_TEMPLATE_DIR / fname
        p = tenant_p if tenant_p.exists() else default_p
        if p.exists():
            stat = p.stat()
            result[mode] = {
                "exists":    True,
                "filename":  fname,
                "size_kb":   round(stat.st_size / 1024, 1),
                "modified":  datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                "scope":     "tenant" if p == tenant_p else "shared_default",
            }
        else:
            result[mode] = {"exists": False}
    return jsonify(result)


# ── Generate Default Templates ──

@app.route("/admin/generate-templates", methods=["POST"])
def admin_generate_templates():
    """기본 템플릿 자동 생성 (openpyxl)"""
    if not _check_admin(request):
        return jsonify({"error": "Invalid admin password."}), 403
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["template.manage"])
    if not ok:
        return jsonify({"error": err}), 403

    body  = request.json or {}
    modes = body.get("modes", ["domestic", "international"])

    try:
        from generate_templates import create_domestic_template, create_international_template
        cfg = load_config(host)
        created = []

        if "domestic" in modes:
            p = _tenant_template_dir(host) / "template_domestic.xlsx"
            create_domestic_template(str(p), cfg)
            created.append("template_domestic.xlsx")

        if "international" in modes:
            p = _tenant_template_dir(host) / "template_international.xlsx"
            create_international_template(str(p), cfg)
            created.append("template_international.xlsx")

        return jsonify({"status": "ok", "created": created})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/template-preview")
def admin_template_preview():
    if not _check_admin(request):
        return jsonify({"error": "Admin authentication required."}), 401
    host = _get_request_host()
    ok, err = _require_any_permission(host, ["template.manage"])
    if not ok:
        return jsonify({"error": err}), 403
    mode = (request.args.get("mode") or "domestic").strip().lower()
    if mode not in {"domestic", "international", "default"}:
        mode = "domestic"
    p = get_template_path(mode, host=host)
    if not p:
        return jsonify({"error": "Template not found."}), 404
    try:
        wb = load_workbook(str(p), data_only=False)
        sheets = []
        for ws in wb.worksheets[:3]:
            rows = []
            max_r = min(ws.max_row or 0, 20)
            max_c = min(ws.max_column or 0, 10)
            for r in range(1, max_r + 1):
                row = []
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    v = cell.value
                    row.append("" if v is None else str(v))
                if any(x.strip() for x in row):
                    rows.append(row)
            sheets.append({"name": ws.title, "rows": rows[:12]})
        wb.close()
        return jsonify(
            {
                "status": "ok",
                "mode": mode,
                "filename": p.name,
                "scope": "tenant" if str(_tenant_template_dir(host)) in str(p) else "shared_default",
                "sheets": sheets,
            }
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    port  = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
