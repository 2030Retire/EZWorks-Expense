"""
Excel Filler Module v2
OCR 결과 데이터를 경비보고서 Excel 템플릿에 자동 입력
- Domestic (USD only) / International (Foreign + USD) 지원
- config.json 기반 동적 Expense Type 지원
"""
import shutil
from datetime import date, datetime, timedelta
from collections import defaultdict
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────────────
DATE_ROW       = 8    # 주간 시트의 날짜 헤더 행 (generate_templates 기준)
DATE_START_COL = 4    # D 열 = 4
DATE_MAX_COLS  = 7    # 최대 7일 (D~J)
DEFAULT_EXCHANGE_RATE = 1.0   # Domestic 은 USD 단일; 실제 사용 시 프론트에서 설정


def parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def get_rate_for_date(d, currency, exchange_rate, exchange_rates):
    c = str(currency or "USD").strip().upper() or "USD"
    if c == "USD":
        return 1.0
    if exchange_rates and d:
        date_str = d.strftime("%Y-%m-%d")
        if date_str in exchange_rates:
            v = exchange_rates[date_str]
            if isinstance(v, dict):
                if c in v and v[c]:
                    return float(v[c])
            elif v:
                return float(v)
    return float(exchange_rate)


def _normalize_active_types(active_types: list) -> list:
    """
    active_types 를 [{"id","label"}] 형태로 정규화.
    - ["LUNCH", "DINNER"] 형태도 허용
    - [{"id":"LUNCH","label":"Lunch"}] 형태도 허용
    """
    normalized = []
    for item in active_types or []:
        if isinstance(item, str):
            type_id = item.strip().upper()
            if type_id:
                normalized.append({"id": type_id, "label": type_id})
            continue
        if isinstance(item, dict) and item.get("id"):
            type_id = str(item["id"]).strip().upper()
            if not type_id:
                continue
            label = str(item.get("label") or type_id).strip()
            normalized.append({"id": type_id, "label": label})
    return normalized


# ──────────────────────────────────────────────────────
# 주간 시트 관리
# ──────────────────────────────────────────────────────

def get_week_end_date(d: date) -> date:
    """날짜가 속한 주의 일요일(Week Ending) 반환."""
    return d + timedelta(days=(6 - d.weekday()) % 7)


def _find_date_row(ws) -> int:
    """
    날짜 헤더 행 번호 찾기.
    generate_templates 는 row 8 (DATE_ROW) 에 날짜를 넣음.
    혹시 다른 위치일 경우에도 동작하도록 스캔.
    """
    for row in range(1, 20):
        for col in range(DATE_START_COL, DATE_START_COL + DATE_MAX_COLS):
            v = ws.cell(row=row, column=col).value
            d = parse_date(v)
            if d:
                return row
    return DATE_ROW


def _find_category_rows(ws, active_types: list, date_row: int) -> dict:
    """
    카테고리 ID → 행 번호 매핑.
    generate_templates 기준: date_row + 1 부터 순서대로 active_types.
    (기존 Daedong USA 양식 같은 커스텀 템플릿의 경우 label로 매칭 시도)
    """
    # 1) label 매칭 (모든 타입 id→label 맵)
    label_to_id = {t["label"].upper(): t["id"] for t in active_types}
    id_to_row   = {}

    for row in range(date_row + 1, date_row + 60):
        v = str(ws.cell(row=row, column=1).value or "").strip()
        if not v:
            continue
        if v.upper() in label_to_id:
            id_to_row[label_to_id[v.upper()]] = row

    # 2) label 매칭이 없으면 순서 기반 (generate_templates 생성 시트)
    if not id_to_row:
        for idx, t in enumerate(active_types):
            id_to_row[t["id"]] = date_row + 1 + idx

    return id_to_row


def auto_create_weekly_sheets(wb, receipt_dates: list,
                               template_sheet_name: str = "Report") -> dict:
    if template_sheet_name not in wb.sheetnames:
        return {}

    template_ws = wb[template_sheet_name]
    week_ends = sorted(set(get_week_end_date(d) for d in receipt_dates if d))

    weekly = {}
    for week_end in week_ends:
        sheet_name = f"{week_end.month:02d}{week_end.day:02d}"

        if sheet_name not in wb.sheetnames:
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = sheet_name
        else:
            new_ws = wb[sheet_name]

        date_row   = _find_date_row(new_ws)
        week_start = week_end - timedelta(days=6)

        # 날짜 헤더 기입
        date_col_map = {}
        d, col = week_start, DATE_START_COL
        while d <= week_end and col < DATE_START_COL + DATE_MAX_COLS:
            new_ws.cell(row=date_row, column=col).value = d
            date_col_map[d] = col
            d += timedelta(days=1)
            col += 1

        weekly[sheet_name] = {
            "ws":           new_ws,
            "date_row":     date_row,
            "date_set":     set(date_col_map.keys()),
            "date_col_map": date_col_map,
        }

    return weekly


def detect_weekly_sheets(wb, year: Optional[int] = None) -> dict:
    if year is None:
        year = datetime.now().year
    weekly = {}
    for name in wb.sheetnames:
        if not (name.isdigit() and len(name) in (3, 4)):
            continue
        ws       = wb[name]
        date_row = _find_date_row(ws)
        date_set = set()
        date_col_map = {}

        for col in range(DATE_START_COL, DATE_START_COL + DATE_MAX_COLS):
            d = parse_date(ws.cell(row=date_row, column=col).value)
            if d:
                date_set.add(d)
                date_col_map[d] = col

        if not date_set:
            # 시트 이름에서 주 범위 추정
            try:
                nm = name.zfill(4)
                month, day = int(nm[:2]), int(nm[2:])
                end_d  = date(year, month, day)
                start_d = end_d - timedelta(days=6)
                d, col = start_d, DATE_START_COL
                while d <= end_d and col < DATE_START_COL + DATE_MAX_COLS:
                    date_set.add(d)
                    date_col_map[d] = col
                    d   += timedelta(days=1)
                    col += 1
            except Exception:
                pass

        weekly[name] = {
            "ws":           ws,
            "date_row":     date_row,
            "date_set":     date_set,
            "date_col_map": date_col_map,
        }
    return weekly


def find_weekly_sheet_for_date(d: date, weekly_sheets: dict) -> Optional[dict]:
    for info in weekly_sheets.values():
        if d in info["date_set"]:
            return info
    return None


# ──────────────────────────────────────────────────────
# List 시트 입력
# ──────────────────────────────────────────────────────

def detect_list_sheet(wb):
    for name in ["List", "list", "LIST", "목록", "리스트"]:
        if name in wb.sheetnames:
            return wb[name]
    return None


def find_list_data_start_row(ws) -> int:
    for row in range(1, 25):
        val = str(ws.cell(row=row, column=1).value or "").strip()
        if val == "1" or val.isdigit():
            return row
    return 9   # generate_templates 기준 데이터 시작 행


def _find_list_account_code_col(ws, start_row: int) -> Optional[int]:
    """
    List 시트에서 계정코드(Account/GL) 열을 헤더 텍스트로 탐지.
    없으면 None 반환하고 기존 열 구조를 그대로 사용한다.
    """
    keywords = ("account", "account code", "gl", "gl code", "ledger", "계정")
    for row in range(1, max(2, start_row + 1)):
        for col in range(1, 16):
            v = str(ws.cell(row=row, column=col).value or "").strip().lower()
            if not v:
                continue
            if any(k in v for k in keywords):
                return col
    return None


def fill_list_sheet_domestic(ws, receipts, start_row: int,
                              exchange_rate: float = 1.0, exchange_rates=None):
    """
    Domestic List 시트:
    A=No, B=Date, C=Merchant, D=Category, E=Amount(USD), F=Memo
    """
    account_col = _find_list_account_code_col(ws, start_row)
    for i, r in enumerate(receipts, 1):
        row     = start_row + i - 1
        r_date  = parse_date(r.get("date"))
        amount  = float(r.get("amount") or 0)
        currency = r.get("currency", "USD")
        rate     = get_rate_for_date(r_date, currency, exchange_rate, exchange_rates)

        if currency == "USD":
            amount_usd = amount
        else:
            amount_usd = amount * rate if rate else 0.0

        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = r_date
        ws.cell(row=row, column=2).number_format = "mm/dd/yyyy"
        ws.cell(row=row, column=3).value = r.get("merchant", "")
        ws.cell(row=row, column=4).value = r.get("category", "MISCELLANEOUS")
        ws.cell(row=row, column=5).value = round(amount_usd, 2)
        memo = str(r.get("memo", "") or "")
        account_code = str(r.get("account_code", "") or "").strip()
        if account_col:
            ws.cell(row=row, column=account_col).value = account_code
        elif account_code and f"AC:{account_code}" not in memo:
            memo = f"{memo} | AC:{account_code}".strip(" |")
        ws.cell(row=row, column=6).value = memo


def fill_list_sheet_international(ws, receipts, start_row: int,
                                   exchange_rate: float = 1.0, exchange_rates=None):
    """
    International List 시트:
    A=No, B=Date, C=Merchant, D=Category,
    E=Currency, F=ForeignAmt, G=ExchRate, H=USD, I=Memo
    """
    account_col = _find_list_account_code_col(ws, start_row)
    for i, r in enumerate(receipts, 1):
        row      = start_row + i - 1
        r_date   = parse_date(r.get("date"))
        amount   = float(r.get("amount") or 0)
        currency = r.get("currency", "USD")
        rate     = get_rate_for_date(r_date, currency, exchange_rate, exchange_rates)

        if currency == "USD":
            amount_usd    = amount
            amount_foreign = amount
            effective_rate = 1.0
        else:
            amount_usd     = amount * rate if rate else 0.0
            amount_foreign = amount
            effective_rate = rate

        ws.cell(row=row, column=1).value = i
        ws.cell(row=row, column=2).value = r_date
        ws.cell(row=row, column=2).number_format = "mm/dd/yyyy"
        ws.cell(row=row, column=3).value = r.get("merchant", "")
        ws.cell(row=row, column=4).value = r.get("category", "MISCELLANEOUS")
        ws.cell(row=row, column=5).value = currency
        ws.cell(row=row, column=6).value = round(amount_foreign, 2)
        ws.cell(row=row, column=7).value = round(effective_rate, 4)
        ws.cell(row=row, column=8).value = round(amount_usd, 2)
        memo = str(r.get("memo", "") or "")
        account_code = str(r.get("account_code", "") or "").strip()
        if account_col:
            ws.cell(row=row, column=account_col).value = account_code
        elif account_code and f"AC:{account_code}" not in memo:
            memo = f"{memo} | AC:{account_code}".strip(" |")
        ws.cell(row=row, column=9).value = memo


# ──────────────────────────────────────────────────────
# 주간 시트 입력
# ──────────────────────────────────────────────────────

def fill_weekly_sheet_domestic(ws, receipts_by_date: dict, date_col_map: dict,
                                active_types: list,
                                exchange_rate: float = 1.0, exchange_rates=None):
    """Domestic 주간 시트 — 날짜열 × 카테고리행 에 USD 금액 입력."""
    cat_row_map = _find_category_rows(ws, active_types, _find_date_row(ws))

    for d, day_receipts in receipts_by_date.items():
        col = date_col_map.get(d)
        if col is None:
            continue
        cat_totals = defaultdict(float)
        for r in day_receipts:
            amount   = float(r.get("amount") or 0)
            currency = r.get("currency", "USD")
            rate     = get_rate_for_date(d, currency, exchange_rate, exchange_rates)
            usd      = amount if currency == "USD" else (amount * rate if rate else 0.0)
            cat_totals[r.get("category", "MISCELLANEOUS")] += usd

        for cat_id, row_idx in cat_row_map.items():
            amt = cat_totals.get(cat_id, 0)
            if amt:
                current = ws.cell(row=row_idx, column=col).value or 0
                ws.cell(row=row_idx, column=col).value = round(current + amt, 2)


def fill_weekly_sheet_international(ws, receipts_by_date: dict, date_col_map: dict,
                                     active_types: list,
                                     exchange_rate: float = 1.0, exchange_rates=None):
    """
    International 주간 시트.
    generate_templates 기준: 각 날짜마다 2열 (Foreign col, USD col).
    날짜 col D, E → Mon Foreign, Mon USD
              F, G → Tue Foreign, Tue USD ...
    """
    date_row    = _find_date_row(ws)
    cat_row_map = _find_category_rows(ws, active_types, date_row)

    # 날짜 → (foreign_col, usd_col) 매핑 구성
    # generate_templates: 7일 × 2열 → C(3)~P(16)
    # date_col_map 의 col 값은 D=4 기준이므로, 국제 양식에선 실제 col = 3 + (day_idx * 2)
    sorted_dates = sorted(date_col_map.keys())
    intl_col_map = {}   # date → (foreign_col, usd_col)
    for day_idx, d in enumerate(sorted_dates):
        foreign_col = 3 + day_idx * 2       # C, E, G, I, K, M, O
        usd_col     = foreign_col + 1
        intl_col_map[d] = (foreign_col, usd_col)

    for d, day_receipts in receipts_by_date.items():
        cols = intl_col_map.get(d)
        if cols is None:
            continue
        foreign_col, usd_col = cols

        cat_foreign = defaultdict(float)
        cat_usd     = defaultdict(float)
        for r in day_receipts:
            amount   = float(r.get("amount") or 0)
            currency = r.get("currency", "USD")
            rate     = get_rate_for_date(d, currency, exchange_rate, exchange_rates)
            if currency == "USD":
                usd = amount; foreign = amount
            else:
                usd = amount * rate if rate else 0.0
                foreign = amount
            cat_id = r.get("category", "MISCELLANEOUS")
            cat_foreign[cat_id] += foreign
            cat_usd[cat_id]     += usd

        for cat_id, row_idx in cat_row_map.items():
            if cat_foreign.get(cat_id, 0):
                cur = ws.cell(row=row_idx, column=foreign_col).value or 0
                ws.cell(row=row_idx, column=foreign_col).value = round(cur + cat_foreign[cat_id], 2)
            if cat_usd.get(cat_id, 0):
                cur = ws.cell(row=row_idx, column=usd_col).value or 0
                ws.cell(row=row_idx, column=usd_col).value = round(cur + cat_usd[cat_id], 2)


# ──────────────────────────────────────────────────────
# 직원 정보 입력
# ──────────────────────────────────────────────────────

def fill_employee_info(wb, employee_info: dict, trip_title: str = "",
                       submission_date=None, settlement_month=None):
    """
    모든 시트의 직원 정보 셀을 채움.
    generate_templates 기준:
    - 배너 아래 row 4~6 에 info_label 셀들이 있고, 바로 오른쪽 셀에 값 입력
    """
    name       = employee_info.get("name", "")
    department = employee_info.get("department", "")
    emp_id     = employee_info.get("employee_id", "")
    manager    = employee_info.get("manager", "")
    project    = employee_info.get("project", "")
    period_str = ""
    if submission_date:
        sd = parse_date(submission_date)
        period_str = sd.strftime("%m/%d/%Y") if sd else str(submission_date)

    label_value_map = {
        "employee name": name,
        "name":          name,
        "department":    department,
        "dept":          department,
        "employee id":   emp_id,
        "manager":       manager,
        "supervisor":    manager,
        "project":       project,
        "cost center":   project,
        "expense period":period_str,
        "period":        period_str,
        "trip":          trip_title,
        "destination":   "",    # 비워둠 — 사용자가 직접 입력
        "purpose":       "",
    }

    for ws in wb.worksheets:
        for col in range(2, 10):
            if ws.cell(row=DATE_ROW, column=col).value is not None:
                ws.column_dimensions[get_column_letter(col)].width = max(
                    float(ws.column_dimensions[get_column_letter(col)].width or 0), 12.0
                )
                ws.cell(row=DATE_ROW, column=col).number_format = "mm/dd/yyyy"
        for row in range(1, 15):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                v    = str(cell.value or "").strip().lower()
                if v in label_value_map:
                    # 오른쪽 옆 셀(또는 병합 블록 첫 셀)에 값 입력
                    val_col = col + 1
                    # 이미 병합된 구간이면 병합 영역 내 첫 셀 찾기
                    target  = ws.cell(row=row, column=val_col)
                    if label_value_map[v]:
                        try:
                            target.value = label_value_map[v]
                        except AttributeError:
                            # MergedCell — find the top-left cell of the merge
                            for rng in ws.merged_cells.ranges:
                                if target.coordinate in rng:
                                    ws.cell(row=rng.min_row, column=rng.min_col).value = label_value_map[v]
                                    break


# ──────────────────────────────────────────────────────
# 메인 진입점
# ──────────────────────────────────────────────────────

def fill_expense_report(
    template_path: str,
    receipts: list,
    output_path: str,
    mode: str = "domestic",                   # "domestic" | "international"
    employee_info: dict = None,
    trip_title: str = "",
    exchange_rate: float = DEFAULT_EXCHANGE_RATE,
    exchange_rates: Optional[dict] = None,
    date_range: Optional[tuple] = None,
    submission_date=None,
    settlement_month=None,
    active_types: list = None,
    company_info: dict = None,
) -> str:
    employee_info = employee_info or {}
    active_types  = _normalize_active_types(active_types or [])
    company_info  = company_info or {}

    # 기본 카테고리 fallback (active_types 가 비어있을 때)
    if not active_types:
        active_types = [
            {"id": "BREAKFAST",     "label": "Breakfast"},
            {"id": "LUNCH",         "label": "Lunch"},
            {"id": "DINNER",        "label": "Dinner"},
            {"id": "ENTERTAINMENT", "label": "Entertainment"},
            {"id": "LODGING",       "label": "Lodging / Hotel"},
            {"id": "AIRFARE",       "label": "Airfare"},
            {"id": "CAR_RENTAL",    "label": "Car Rental"},
            {"id": "TAXI",          "label": "Taxi / Rideshare"},
            {"id": "MILEAGE",       "label": "Mileage"},
            {"id": "PARKING",       "label": "Parking / Tolls"},
            {"id": "PHONE",         "label": "Phone / Internet"},
            {"id": "MISCELLANEOUS", "label": "Miscellaneous"},
        ]

    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)

    # ── 정렬 + 필터 ──
    def sort_key(r):
        return parse_date(r.get("date")) or date.min

    receipts_sorted = sorted(receipts, key=sort_key)
    if date_range:
        start_d, end_d = date_range
        receipts_sorted = [
            r for r in receipts_sorted
            if (lambda d: d and start_d <= d <= end_d)(parse_date(r.get("date")))
        ]

    receipt_dates = [parse_date(r.get("date")) for r in receipts_sorted if r.get("date")]
    year          = receipt_dates[0].year if receipt_dates else datetime.now().year

    # ── List 시트 ──
    ws_list = detect_list_sheet(wb)
    if ws_list:
        start_row = find_list_data_start_row(ws_list)
        if mode == "international":
            fill_list_sheet_international(ws_list, receipts_sorted, start_row,
                                          exchange_rate, exchange_rates)
        else:
            fill_list_sheet_domestic(ws_list, receipts_sorted, start_row,
                                     exchange_rate, exchange_rates)

    # ── 주간 시트 ──
    weekly_sheets = detect_weekly_sheets(wb, year=year)
    if not weekly_sheets and receipt_dates:
        weekly_sheets = auto_create_weekly_sheets(wb, receipt_dates)

    # 날짜별 영수증 그룹핑
    sheet_receipt_map = defaultdict(lambda: defaultdict(list))
    for r in receipts_sorted:
        d = parse_date(r.get("date"))
        if not d:
            continue
        info = find_weekly_sheet_for_date(d, weekly_sheets)
        if info:
            sheet_receipt_map[info["ws"].title][d].append(r)

    for sheet_name, receipts_by_date in sheet_receipt_map.items():
        info = weekly_sheets[sheet_name]
        if mode == "international":
            fill_weekly_sheet_international(
                info["ws"], receipts_by_date, info["date_col_map"],
                active_types, exchange_rate, exchange_rates
            )
        else:
            fill_weekly_sheet_domestic(
                info["ws"], receipts_by_date, info["date_col_map"],
                active_types, exchange_rate, exchange_rates
            )

    # ── 직원 정보 ──
    fill_employee_info(wb, employee_info, trip_title, submission_date, settlement_month)

    wb.save(output_path)
    return output_path
